"""
Router de Propiedades - Endpoints de búsqueda y filtrado de propiedades
"""
from fastapi import APIRouter, Query
from typing import List, Optional
from sqlmodel import Session, select, func
from sqlalchemy import and_, or_
from config.db_connection import engine
from models.property import Property
from models.city import City
from services.stats_service import get_local_now
from services.property_filters import (
    build_rooms_filter, build_baths_filter, build_garages_filter,
    build_stratum_filter, build_antiquity_filter, build_property_type_filter,
    build_price_type_filters, format_antiquity
)
from services.geo_service import geocode_address, filter_properties_by_distance
import re

router = APIRouter(prefix="/api", tags=["properties"])


def format_property_data(prop, city, distance_map: dict = None):
    """Formatear datos de propiedad para respuesta"""
    # Link de FincaRaiz
    finca_raiz_link = None
    if hasattr(prop, 'fr_property_id') and prop.fr_property_id:
        finca_raiz_link = f"https://www.fincaraiz.com.co/inmueble/{prop.fr_property_id}"
    
    # Link de Google Maps
    maps_link = None
    if prop.latitude and prop.longitude:
        maps_link = f"https://www.google.com/maps?q={prop.latitude},{prop.longitude}"
    
    # Procesar stratum
    stratum_value = None
    if prop.stratum:
        stratum_match = re.search(r'Estrato\s*(\d+)', prop.stratum)
        if stratum_match:
            stratum_value = int(stratum_match.group(1))
    
    # Procesar antiquity
    antiquity_display = format_antiquity(prop.antiquity)
    
    # Convertir rooms, baths, garages a números
    rooms_value = None
    if prop.rooms:
        try:
            rooms_value = int(prop.rooms) if isinstance(prop.rooms, str) else prop.rooms
        except:
            rooms_value = prop.rooms
    
    baths_value = None
    if prop.baths:
        try:
            baths_value = int(prop.baths) if isinstance(prop.baths, str) else prop.baths
        except:
            baths_value = prop.baths
    
    garages_value = None
    if prop.garages:
        try:
            garages_value = int(prop.garages) if isinstance(prop.garages, str) else prop.garages
        except:
            garages_value = prop.garages
    
    return {
        "id": prop.fr_property_id,
        "city": city.name if city else "Sin especificar",
        "area": prop.area,
        "rooms": rooms_value,
        "price": prop.price,
        "offer_type": "Venta" if prop.offer == "sell" else "Renta",
        "creation_date": prop.creation_date.isoformat() if prop.creation_date else None,
        "last_update": prop.last_update.isoformat() if prop.last_update else None,
        "title": prop.title,
        "finca_raiz_link": finca_raiz_link,
        "maps_link": maps_link,
        "latitude": prop.latitude,
        "longitude": prop.longitude,
        "baths": baths_value,
        "garages": garages_value,
        "stratum": stratum_value,
        "antiquity": antiquity_display,
        "is_new": prop.is_new,
        "address": getattr(prop, 'address', None),
        "distance": distance_map.get(prop.fr_property_id, None) if distance_map else None
    }


@router.get("/properties")
async def get_properties(
    page: int = 1,
    limit: int = 50,
    city_ids: Optional[List[int]] = Query(None),
    offer_type: str = None,
    min_price: float = None,
    max_price: float = None,
    min_area: float = None,
    max_area: float = None,
    rooms: Optional[List[str]] = Query(None),
    baths: Optional[List[str]] = Query(None),
    garages: Optional[List[str]] = Query(None),
    stratums: Optional[List[str]] = Query(None),
    antiquity_categories: Optional[List[int]] = Query(None),
    antiquity_filter: str = None,
    property_type: Optional[List[str]] = Query(None),
    min_sale_price: float = None,
    max_sale_price: float = None,
    min_rent_price: float = None,
    max_rent_price: float = None,
    updated_date_from: str = None,
    updated_date_to: str = None,
    search_address: str = None,
    latitude: float = None,
    longitude: float = None,
    radius: int = None
):
    """Obtener propiedades con filtros y paginación"""
    try:
        with Session(engine) as session:
            query = select(Property, City).outerjoin(City, Property.city_id == City.id)
            filters = []
            
            # Aplicar filtros básicos
            if city_ids:
                filters.append(Property.city_id.in_(city_ids))
            
            if offer_type:
                filters.append(Property.offer == offer_type)
                
            if min_price is not None:
                filters.append(Property.price >= min_price)
                
            if max_price is not None:
                filters.append(Property.price <= max_price)
            
            # Filtros de precio por tipo de oferta
            price_filter = build_price_type_filters(Property, min_sale_price, max_sale_price, min_rent_price, max_rent_price)
            if price_filter is not None:
                filters.append(price_filter)
                
            if min_area is not None:
                filters.append(Property.area >= min_area)
                
            if max_area is not None:
                filters.append(Property.area <= max_area)
            
            # Filtros construidos con servicios
            rooms_filter = build_rooms_filter(Property, rooms)
            if rooms_filter is not None:
                filters.append(rooms_filter)
            
            baths_filter = build_baths_filter(Property, baths)
            if baths_filter is not None:
                filters.append(baths_filter)
            
            garages_filter = build_garages_filter(Property, garages)
            if garages_filter is not None:
                filters.append(garages_filter)
            
            stratum_filter = build_stratum_filter(Property, stratums)
            if stratum_filter is not None:
                filters.append(stratum_filter)
            
            antiquity_filter_result = build_antiquity_filter(Property, antiquity_categories, antiquity_filter)
            if antiquity_filter_result is not None:
                filters.append(antiquity_filter_result)
            
            property_type_filter = build_property_type_filter(Property, property_type)
            if property_type_filter is not None:
                filters.append(property_type_filter)
            
            # Filtros de fecha
            if updated_date_from:
                from datetime import datetime
                try:
                    date_from = datetime.strptime(updated_date_from, "%Y-%m-%d").date()
                    filters.append(Property.last_update >= date_from)
                except:
                    pass
            
            if updated_date_to:
                from datetime import datetime
                try:
                    date_to = datetime.strptime(updated_date_to, "%Y-%m-%d").date()
                    filters.append(Property.last_update <= date_to)
                except:
                    pass
            
            # Geocodificar dirección si es necesario
            if search_address and radius is not None:
                lat, lng, _ = geocode_address(search_address)
                if lat and lng:
                    latitude, longitude = lat, lng
            
            distance_map = {}
            
            # Filtrar por distancia si hay coordenadas
            if latitude is not None and longitude is not None and radius is not None:
                if filters:
                    query = query.where(and_(*filters))
                
                all_results = session.exec(query.order_by(Property.creation_date.desc())).all()
                
                filtered_results, distance_map = filter_properties_by_distance(all_results, latitude, longitude, radius)
                total_count = len(filtered_results)
                
                # Paginación después del filtrado
                offset = (page - 1) * limit
                results = filtered_results[offset:offset + limit]
            else:
                # Sin filtro de distancia
                count_query = select(func.count(Property.fr_property_id)).outerjoin(City, Property.city_id == City.id)
                if filters:
                    count_query = count_query.where(and_(*filters))
                
                total_count = session.exec(count_query).one()
                
                if filters:
                    query = query.where(and_(*filters))
                
                offset = (page - 1) * limit
                query = query.offset(offset).limit(limit).order_by(Property.creation_date.desc())
                results = session.exec(query).all()
            
            # Formatear resultados
            properties = [format_property_data(prop, city, distance_map) for prop, city in results]
            
            total_pages = (total_count + limit - 1) // limit
            
            return {
                "status": "success",
                "data": {
                    "properties": properties,
                    "pagination": {
                        "page": page,
                        "limit": limit,
                        "total_count": total_count,
                        "total_pages": total_pages,
                        "has_next": page < total_pages,
                        "has_prev": page > 1
                    }
                }
            }
    except Exception as e:
        return {"status": "error", "detail": str(e)}


@router.get("/properties/by-zone")
async def get_properties_by_zone(
    boundary_type: str = Query(..., description="Tipo de límite"),
    city_id: Optional[int] = Query(None),
    north: Optional[float] = Query(None),
    south: Optional[float] = Query(None),
    east: Optional[float] = Query(None),
    west: Optional[float] = Query(None),
    property_type: Optional[str] = Query(None),
    updated_date_from: Optional[str] = Query(None),
    updated_date_to: Optional[str] = Query(None)
):
    """Obtener propiedades agrupadas por zona"""
    try:
        with Session(engine) as session:
            filters = [
                Property.latitude.isnot(None),
                Property.longitude.isnot(None)
            ]
            
            if city_id:
                filters.append(Property.city_id == city_id)
            
            if all([north, south, east, west]):
                filters.extend([
                    Property.latitude <= north,
                    Property.latitude >= south,
                    Property.longitude <= east,
                    Property.longitude >= west
                ])
            
            # Filtrar por tipo de propiedad
            if property_type:
                type_filter = build_property_type_filter(Property, property_type.split(','))
                if type_filter is not None:
                    filters.append(type_filter)
            
            # Filtrar por fechas
            if updated_date_from:
                from datetime import datetime
                date_from = datetime.strptime(updated_date_from, '%Y-%m-%d').date()
                filters.append(Property.last_update >= date_from)
            
            if updated_date_to:
                from datetime import datetime
                date_to = datetime.strptime(updated_date_to, '%Y-%m-%d').date()
                filters.append(Property.last_update <= date_to)
            
            query = select(Property).where(and_(*filters))
            properties = session.exec(query).all()
            
            properties_data = [{
                'id': prop.fr_property_id,
                'latitude': prop.latitude,
                'longitude': prop.longitude,
                'price': prop.price,
                'offer': prop.offer,
                'area': prop.area,
                'rooms': prop.rooms,
                'city_id': prop.city_id,
                'location_main': prop.location_main,
                'stratum': prop.stratum,
                'title': prop.title,
                'last_update': prop.last_update.isoformat() if prop.last_update else None
            } for prop in properties]
            
            properties_for_sale = [p for p in properties_data if p['offer'] == 'sell']
            properties_for_rent = [p for p in properties_data if p['offer'] == 'rent']
            
            return {
                'status': 'success',
                'boundary_type': boundary_type,
                'data': {
                    'properties': properties_data,
                    'summary': {
                        'total': len(properties_data),
                        'for_sale': len(properties_for_sale),
                        'for_rent': len(properties_for_rent)
                    }
                }
            }
            
    except Exception as e:
        print(f"Error getting properties by zone: {e}")
        return {'status': 'error', 'message': str(e), 'data': None}


@router.post("/properties/send-excel")
async def send_properties_excel(request: dict):
    """Enviar propiedades por email en formato Excel"""
    import smtplib
    import os
    import io
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlalchemy import cast, Integer
    from math import radians, cos, sin, asin, sqrt
    
    try:
        # Obtener parámetros del request
        email_destinatario = request.get('email')
        filters = request.get('filters', {})
        
        print(f"📧 Solicitud de Excel para: {email_destinatario}")
        print(f"🔍 Filtros recibidos: {filters}")
        
        if not email_destinatario:
            return {"status": "error", "detail": "Email es requerido"}
        
        with Session(engine) as session:
            query = select(Property, City).outerjoin(City, Property.city_id == City.id)
            
            # Aplicar filtros
            filter_conditions = []
            
            if filters.get('city_ids'):
                filter_conditions.append(Property.city_id.in_(filters['city_ids']))
            
            if filters.get('offer_type'):
                filter_conditions.append(Property.offer == filters['offer_type'])
                
            if filters.get('min_price') is not None:
                filter_conditions.append(Property.price >= filters['min_price'])
                
            if filters.get('max_price') is not None:
                filter_conditions.append(Property.price <= filters['max_price'])

            # Filtros de precio específicos por tipo de oferta
            sale_price_conditions = []
            rent_price_conditions = []
            
            if filters.get('min_sale_price') is not None:
                sale_price_conditions.append(Property.price >= filters['min_sale_price'])
            if filters.get('max_sale_price') is not None:
                sale_price_conditions.append(Property.price <= filters['max_sale_price'])
                
            if filters.get('min_rent_price') is not None:
                rent_price_conditions.append(Property.price >= filters['min_rent_price'])
            if filters.get('max_rent_price') is not None:
                rent_price_conditions.append(Property.price <= filters['max_rent_price'])
            
            # Aplicar filtros independientes por tipo de oferta
            price_type_conditions = []
            if sale_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'sell',
                    *sale_price_conditions
                ))
            if rent_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'rent', 
                    *rent_price_conditions
                ))
            
            if price_type_conditions:
                filter_conditions.append(or_(*price_type_conditions))
                
            if filters.get('min_area') is not None:
                filter_conditions.append(Property.area >= filters['min_area'])
                
            if filters.get('max_area') is not None:
                filter_conditions.append(Property.area <= filters['max_area'])
            
            # Filtro de habitaciones
            if filters.get('rooms') is not None and len(filters['rooms']) > 0:
                rooms_conditions = []
                for room in filters['rooms']:
                    if room == "unspecified":
                        rooms_conditions.append(or_(
                            Property.rooms == None,
                            Property.rooms == '',
                            Property.rooms == 'N/A',
                            Property.rooms == 'Sin especificar'
                        ))
                    elif room.endswith('+'):
                        min_value = int(room[:-1])
                        rooms_conditions.append(and_(
                            Property.rooms.regexp_match('^[0-9]+$'),
                            cast(Property.rooms, Integer) >= min_value
                        ))
                    else:
                        rooms_conditions.append(Property.rooms == room)
                if rooms_conditions:
                    filter_conditions.append(or_(*rooms_conditions))
            
            # Filtro de baños
            if filters.get('baths') is not None and len(filters['baths']) > 0:
                baths_conditions = []
                for bath in filters['baths']:
                    if bath == "unspecified":
                        baths_conditions.append(or_(
                            Property.baths == None,
                            Property.baths == '',
                            Property.baths == 'N/A',
                            Property.baths == 'Sin especificar'
                        ))
                    elif bath.endswith('+'):
                        min_value = int(bath[:-1])
                        baths_conditions.append(and_(
                            Property.baths.regexp_match('^[0-9]+$'),
                            cast(Property.baths, Integer) >= min_value
                        ))
                    else:
                        baths_conditions.append(Property.baths == bath)
                if baths_conditions:
                    filter_conditions.append(or_(*baths_conditions))
            
            # Filtro de garajes
            if filters.get('garages') is not None and len(filters['garages']) > 0:
                garages_conditions = []
                for garage in filters['garages']:
                    if garage == "unspecified":
                        garages_conditions.append(or_(
                            Property.garages == None,
                            Property.garages == '',
                            Property.garages == 'N/A',
                            Property.garages == 'Sin especificar'
                        ))
                    elif garage.endswith('+'):
                        min_value = int(garage[:-1])
                        garages_conditions.append(and_(
                            Property.garages.regexp_match('^[0-9]+$'),
                            cast(Property.garages, Integer) >= min_value
                        ))
                    else:
                        garages_conditions.append(Property.garages == garage)
                if garages_conditions:
                    filter_conditions.append(or_(*garages_conditions))
            
            # Filtro de estrato
            if filters.get('stratums') is not None and len(filters['stratums']) > 0:
                stratum_conditions = []
                for stratum in filters['stratums']:
                    if stratum == "unspecified":
                        stratum_conditions.append(or_(
                            Property.stratum == None,
                            Property.stratum == '',
                            Property.stratum == 'Sin especificar'
                        ))
                    else:
                        stratum_str = f"Estrato {stratum}"
                        stratum_conditions.append(Property.stratum == stratum_str)
                if stratum_conditions:
                    filter_conditions.append(or_(*stratum_conditions))
            
            # Filtro de antigüedad
            antiquity_categories = filters.get('antiquity_categories')
            antiquity_filter = filters.get('antiquity_filter')
            
            if (antiquity_categories is not None and len(antiquity_categories) > 0) or antiquity_filter == 'unspecified':
                antiquity_conditions = []
                
                if antiquity_categories is not None and len(antiquity_categories) > 0:
                    try:
                        antiquity_categories = [int(x) for x in antiquity_categories]
                    except:
                        pass
                        
                    for antiquity_category in antiquity_categories:
                        if antiquity_category == 1:
                            antiquity_conditions.extend([
                                Property.antiquity == 'LESS_THAN_1_YEAR',
                                Property.antiquity == 'Menos de 1 año',
                                Property.antiquity == '1'
                            ])
                        elif antiquity_category == 2:
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_1_TO_8_YEARS',
                                Property.antiquity == '1 a 8 años',
                                Property.antiquity == '2'
                            ])
                        elif antiquity_category == 3:
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_9_TO_15_YEARS',
                                Property.antiquity == '9 a 15 años',
                                Property.antiquity == '3'
                            ])
                        elif antiquity_category == 4:
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_16_TO_30_YEARS',
                                Property.antiquity == '16 a 30 años',
                                Property.antiquity == '4'
                            ])
                        elif antiquity_category == 5:
                            antiquity_conditions.extend([
                                Property.antiquity == 'MORE_THAN_30_YEARS',
                                Property.antiquity == 'Más de 30 años',
                                Property.antiquity == '5'
                            ])
                
                if antiquity_filter == 'unspecified':
                    antiquity_conditions.extend([
                        Property.antiquity == 'UNDEFINED',
                        Property.antiquity == 'Sin especificar',
                        Property.antiquity == None,
                        Property.antiquity == ''
                    ])
                
                if antiquity_conditions:
                    filter_conditions.append(or_(*antiquity_conditions))
            
            # Filtro de tipo de propiedad
            property_type = filters.get('property_type')
            if property_type and len(property_type) > 0:
                type_filter = build_property_type_filter(Property, property_type)
                if type_filter is not None:
                    filter_conditions.append(type_filter)
            
            # Filtro de fechas
            if filters.get('updated_date_from'):
                date_from = datetime.strptime(filters.get('updated_date_from'), '%Y-%m-%d').date()
                filter_conditions.append(Property.last_update >= date_from)
            
            if filters.get('updated_date_to'):
                date_to = datetime.strptime(filters.get('updated_date_to'), '%Y-%m-%d').date()
                filter_conditions.append(Property.last_update <= date_to)
            
            if filter_conditions:
                query = query.where(and_(*filter_conditions))
            
            # Obtener propiedades
            results = session.exec(query.order_by(Property.creation_date.desc())).all()
            print(f"📊 Propiedades encontradas antes de filtrar por distancia: {len(results)}")
            
            # Procesar filtro de distancia
            distance_map = {}
            search_lat = filters.get('latitude')
            search_lng = filters.get('longitude')
            radius = filters.get('radius')
            
            if search_lat is not None and search_lng is not None and radius is not None:
                try:
                    search_lat = float(search_lat)
                    search_lng = float(search_lng)
                    radius = float(radius)
                except (ValueError, TypeError) as e:
                    print(f"❌ Error convirtiendo coordenadas: {e}")
                
                print(f"🔍 Procesando filtro de distancia: lat={search_lat}, lng={search_lng}, radius={radius}m")
                
                def calculate_distance(lat1, lng1, lat2, lng2):
                    """Calcula distancia entre dos puntos en metros usando Haversine"""
                    try:
                        R = 6371000
                        lat1, lng1, lat2, lng2 = map(radians, [float(lat1), float(lng1), float(lat2), float(lng2)])
                        dlat = lat2 - lat1
                        dlng = lng2 - lng1
                        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlng/2)**2
                        c = 2 * asin(sqrt(a))
                        return round(R * c)
                    except Exception as e:
                        return float('inf')
                
                filtered_results = []
                for prop, city in results:
                    if prop.latitude and prop.longitude:
                        distance = calculate_distance(search_lat, search_lng, prop.latitude, prop.longitude)
                        if distance <= radius:
                            distance_map[prop.fr_property_id] = distance
                            filtered_results.append((prop, city))
                
                filtered_results.sort(key=lambda x: distance_map.get(x[0].fr_property_id, float('inf')))
                results = filtered_results
                print(f"✅ Propiedades filtradas por distancia: {len(results)}")
            
            # Crear archivo Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Propiedades"
            
            headers = [
                'ID', 'Título', 'Ciudad', 'Tipo', 'Precio (COP)', 
                'Área (m²)', 'Habitaciones', 'Baños', 'Garajes', 
                'Estrato', 'Antigüedad', 'Distancia (m)', 'Fecha Creación', 
                'Última Actualización', 'FincaRaiz', 'Google Maps'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if len(results) > 50000:
                return {"status": "error", "detail": f"Demasiadas propiedades ({len(results):,}). El límite para Excel es 50,000."}
            
            print(f"📊 Generando Excel con {len(results):,} propiedades...")
            
            for i, (prop, city) in enumerate(results):
                row_idx = i + 2
                ws.cell(row=row_idx, column=1, value=prop.fr_property_id)
                ws.cell(row=row_idx, column=2, value=prop.title)
                ws.cell(row=row_idx, column=3, value=city.name if city else "Sin especificar")
                ws.cell(row=row_idx, column=4, value="Venta" if prop.offer == "sell" else "Renta")
                ws.cell(row=row_idx, column=5, value=prop.price)
                ws.cell(row=row_idx, column=6, value=prop.area)
                ws.cell(row=row_idx, column=7, value=prop.rooms)
                ws.cell(row=row_idx, column=8, value=prop.baths)
                ws.cell(row=row_idx, column=9, value=prop.garages)
                ws.cell(row=row_idx, column=10, value=prop.stratum)
                ws.cell(row=row_idx, column=11, value=format_antiquity(prop.antiquity))
                
                distance = distance_map.get(prop.fr_property_id, None)
                ws.cell(row=row_idx, column=12, value=f"{distance:,} m" if distance is not None else "")
                
                ws.cell(row=row_idx, column=13, value=prop.creation_date.isoformat() if prop.creation_date else "")
                ws.cell(row=row_idx, column=14, value=prop.last_update.isoformat() if prop.last_update else "")
                
                if prop.fr_property_id:
                    cell = ws.cell(row=row_idx, column=15)
                    cell.hyperlink = f"https://www.fincaraiz.com.co/inmueble/{prop.fr_property_id}"
                    cell.value = "Ver en FincaRaiz"
                    cell.style = "Hyperlink"
                
                if prop.latitude and prop.longitude:
                    cell = ws.cell(row=row_idx, column=16)
                    cell.hyperlink = f"https://www.google.com/maps?q={prop.latitude},{prop.longitude}"
                    cell.value = "Ver en Maps"
                    cell.style = "Hyperlink"
            
            # Ajustar ancho de columnas
            column_widths = {'A': 12, 'B': 40, 'C': 15, 'D': 10, 'E': 15, 'F': 12, 'G': 12, 'H': 8, 'I': 8, 'J': 8, 'K': 15, 'L': 15, 'M': 12, 'N': 12, 'O': 15, 'P': 15}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Guardar Excel en memoria
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Enviar email
            print(f"📧 Enviando email a {email_destinatario}...")
            
            smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
            smtp_port = int(os.getenv("SMTP_PORT", "587"))
            smtp_user = os.getenv("SMTP_USER")
            smtp_password = os.getenv("SMTP_PASSWORD")
            from_email = os.getenv("FROM_EMAIL", smtp_user)
            
            if not smtp_user or not smtp_password:
                return {"status": "error", "detail": "Credenciales SMTP no configuradas"}
            
            msg = MIMEMultipart()
            msg['From'] = from_email
            msg['To'] = email_destinatario
            msg['Cc'] = from_email
            msg['Subject'] = f"Dashboard Scraper - Propiedades Exportadas"
            
            address_info = ""
            if search_lat is not None and search_lng is not None and filters.get('search_address'):
                address_info = f"\n• Propiedad consultada: {filters.get('search_address')}"
                if radius:
                    address_info += f"\n• Radio de búsqueda: {radius:,} metros"
            
            body = f"""Dashboard Scraper - Propiedades

Adjunto encontrarás el archivo Excel con las propiedades solicitadas.

Resumen:
• Total de propiedades exportadas: {len(results)}
• Fecha de exportación: {get_local_now().strftime("%d/%m/%Y %H:%M")}{address_info}

El archivo Excel contiene información detallada de cada propiedad incluyendo:
• Información básica (título, ciudad, tipo)
• Detalles de precio y características  
• Enlaces directos a FincaRaiz
• Coordenadas geográficas

Saludos cordiales,
Equipo de Avalúos"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            filename = f"propiedades_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.set_payload(excel_file.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)
            
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
            
            print(f"✅ Email enviado exitosamente a {email_destinatario}")
            
            return {
                "status": "success",
                "message": f"Excel enviado exitosamente a {email_destinatario}",
                "properties_count": len(results)
            }
            
    except Exception as e:
        print(f"❌ Error enviando email: {e}")
        import traceback
        traceback.print_exc()
        return {"status": "error", "detail": str(e)}
