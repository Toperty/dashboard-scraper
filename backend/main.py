"""
Backend que FUNCIONA - Sin problemas de tablas
"""

from datetime import datetime, timedelta
from fastapi import FastAPI, Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
import os
import pytz
from pydantic import BaseModel

app = FastAPI(title="Dashboard API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Import models to register with SQLModel
from models.valuation import Valuation

# Initialize database tables on startup
from config.db_connection import init_db
init_db()
print("✅ Database tables initialized")

# Timezone de Colombia
COLOMBIA_TZ = pytz.timezone('America/Bogota')

def get_local_now():
    """Obtener la hora actual en timezone de Colombia"""
    return datetime.now(COLOMBIA_TZ)

def get_city_status(city):
    """Determinar estado de ciudad basado en offsets y updated"""
    if city.updated:
        return "completed"
    elif city.current_sell_offset != 0 or city.current_rent_offset != 0:
        return "en_proceso"
    else:
        return "programado"


# Modelo Pydantic para el request de avalúo
class PropertyValuationRequest(BaseModel):
    area: float
    rooms: int
    baths: int
    garages: int
    stratum: int
    latitude: float
    longitude: float
    antiquity: int
    is_new: str
    area_per_room: float
    age_bucket: str
    has_garage: int
    city_id: str
    property_type: int

# Modelo Pydantic para guardar avalúo
class SaveValuationRequest(BaseModel):
    valuation_name: str
    area: float
    property_type: int
    rooms: int
    baths: int
    garages: int
    stratum: int
    antiquity: int
    latitude: float
    longitude: float
    capitalization_rate: Optional[float] = None
    sell_price_per_sqm: Optional[float] = None
    rent_price_per_sqm: Optional[float] = None
    total_sell_price: Optional[float] = None
    total_rent_price: Optional[float] = None
    final_price: Optional[float] = None

def get_recent_logs(session):
    """Obtener logs recientes de la tabla scraper_logs"""
    from sqlmodel import select
    from models.scraper_log import ScraperLog
    
    try:
        # Obtener los 10 logs más recientes
        recent_logs_query = select(ScraperLog).order_by(ScraperLog.timestamp.desc()).limit(10)
        recent_logs = session.exec(recent_logs_query).all()
        
        return [
            {
                "id": log.id,
                "timestamp": log.timestamp.isoformat() if log.timestamp else get_local_now().isoformat(),
                "scraper_name": log.scraper_name,
                "city_code": log.city_code,
                "offer_type": log.offer_type,
                "page_number": log.page_number,
                "log_level": log.log_level,
                "log_type": log.log_type,
                "message": log.message,
                "execution_time_ms": log.execution_time_ms,
                "properties_found": log.properties_found,
                "properties_validated": log.properties_validated,
                "error_type": log.error_type,
                "session_id": log.session_id,
                "scheduled_time": log.scheduled_time.isoformat() if log.scheduled_time else None
            }
            for log in recent_logs
        ]
    except Exception as e:
        return []

def get_next_executions(session):
    """Obtener próximas ejecuciones basadas en lógica de orden alfabético"""
    from sqlalchemy import text
    
    try:
        now = get_local_now()
        
        # Determinar cuál es la ciudad que se está ejecutando actualmente
        current_city_query = text("""
            SELECT city_code, offer_type FROM scraper_logs 
            ORDER BY timestamp DESC LIMIT 1
        """)
        
        current_result = session.exec(current_city_query).first()
        current_city_code = current_result[0] if current_result else None
        current_offer_type = current_result[1] if current_result else None
        
        # Obtener información de ciudades y su progreso
        cities_query = text("""
            SELECT 
                c.name,
                c.website_name,
                c.current_sell_offset,
                c.sell_pages_limit,
                c.current_rent_offset,
                c.rent_pages_limit,
                c.updated
            FROM city c
            WHERE c.updated = false
            ORDER BY c.name ASC
        """)
        
        cities_result = session.exec(cities_query).fetchall()
        
        executions = []
        
        # Encontrar la próxima ejecución basada en la lógica:
        # 1. Primero sell, luego rent
        # 2. Orden alfabético
        # 3. Horario cada 30 minutos (:30)
        
        next_execution_time = now.replace(second=0, microsecond=0)
        # Redondear al próximo :30
        if next_execution_time.minute < 30:
            next_execution_time = next_execution_time.replace(minute=30)
        else:
            next_execution_time = next_execution_time.replace(minute=30) + timedelta(hours=1)
        
        execution_count = 0
        for i, city_row in enumerate(cities_result):
            if execution_count >= 5:  # Límite de 5 ejecuciones
                break
                
            name, website_name, sell_offset, sell_limit, rent_offset, rent_limit, updated = city_row
            
            # Determinar si sell está completado
            sell_completed = sell_offset >= sell_limit if sell_limit > 0 else True
            rent_completed = rent_offset >= rent_limit if rent_limit > 0 else True
            
            # Si la ciudad está en progreso actual, determinar qué sigue
            if website_name == current_city_code:
                # Si estamos haciendo sell y no está completado, continuar sell
                if current_offer_type == "sell" and not sell_completed:
                    executions.append({
                        "city": name,
                        "type": "sell",
                        "scheduled_time": next_execution_time.isoformat(),
                        "minutes_remaining": int((next_execution_time - now).total_seconds() / 60)
                    })
                    next_execution_time += timedelta(hours=1)
                    execution_count += 1
                # Si sell está completado, hacer rent
                elif sell_completed and not rent_completed:
                    executions.append({
                        "city": name,
                        "type": "rent",
                        "scheduled_time": next_execution_time.isoformat(),
                        "minutes_remaining": int((next_execution_time - now).total_seconds() / 60)
                    })
                    next_execution_time += timedelta(hours=1)
                    execution_count += 1
                continue
            
            # Para otras ciudades, agregar sell primero si no está completado
            if not sell_completed:
                executions.append({
                    "city": name,
                    "type": "sell",
                    "scheduled_time": next_execution_time.isoformat(),
                    "minutes_remaining": int((next_execution_time - now).total_seconds() / 60)
                })
                next_execution_time += timedelta(hours=1)
                execution_count += 1
                
                if execution_count >= 5:
                    break
            
            # Luego rent si sell está completado y rent no
            if sell_completed and not rent_completed:
                executions.append({
                    "city": name,
                    "type": "rent",
                    "scheduled_time": next_execution_time.isoformat(),
                    "minutes_remaining": int((next_execution_time - now).total_seconds() / 60)
                })
                next_execution_time += timedelta(hours=1)
                execution_count += 1
        
        # Si no hay ejecuciones, mostrar una por defecto
        if not executions:
            executions.append({
                "city": "Sistema",
                "type": "info",
                "scheduled_time": next_execution_time.isoformat(),
                "minutes_remaining": int((next_execution_time - now).total_seconds() / 60)
            })
        
        return executions
        
    except Exception as e:
        # Fallback con próxima ejecución estimada
        now = get_local_now()
        next_time = now.replace(second=0, microsecond=0)
        if next_time.minute < 30:
            next_time = next_time.replace(minute=30)
        else:
            next_time = next_time.replace(minute=30) + timedelta(hours=1)
            
        return [{
            "city": "Sistema",
            "type": "info",
            "scheduled_time": next_time.isoformat(),
            "minutes_remaining": int((next_time - now).total_seconds() / 60)
        }]

def get_property_stats(session, city_id=None):
    """Obtener estadísticas reales de propiedades"""
    from sqlmodel import select, func
    from models.property import Property
    from datetime import date
    
    try:
        today = get_local_now().date()
        
        if city_id:
            # Stats para una ciudad específica
            total_query = select(func.count(Property.fr_property_id)).where(Property.city_id == city_id)
            today_query = select(func.count(Property.fr_property_id)).where(
                Property.city_id == city_id,
                Property.creation_date == today
            )
        else:
            # Stats globales
            total_query = select(func.count(Property.fr_property_id))
            today_query = select(func.count(Property.fr_property_id)).where(Property.creation_date == today)
        
        total_properties = session.exec(total_query).first() or 0
        today_properties = session.exec(today_query).first() or 0
        
        return total_properties, today_properties
    except Exception as e:
        return 0, 0

def get_avg_speed(session):
    """Calcular páginas por minuto desde PAGE_NAVIGATION logs"""
    from sqlmodel import select, func
    from models.scraper_log import ScraperLog
    from datetime import datetime, timedelta
    
    try:
        # Últimas 24 horas
        yesterday = get_local_now() - timedelta(hours=24)
        
        # Contar logs de PAGE_NAVIGATION en las últimas 24 horas
        page_count_query = select(func.count(ScraperLog.id)).where(
            ScraperLog.timestamp >= yesterday,
            ScraperLog.log_type == "PAGE_NAVIGATION"
        )
        
        page_count = session.exec(page_count_query).first()
        
        if page_count and page_count > 0:
            # Calcular páginas por minuto (24 horas = 1440 minutos)
            pages_per_minute = page_count / 1440
            return round(float(pages_per_minute), 2)
        else:
            return 0.0
    except Exception as e:
        print(f"Error calculating pages per minute: {e}")
        return 0.0

def get_last_execution_time(session):
    """Obtener tiempo desde la última ejecución obteniendo el timestamp más reciente de la BD"""
    try:
        from sqlalchemy import text
        
        # Obtener el último timestamp de scraper_logs
        query = text("""
            SELECT MAX(timestamp) FROM scraper_logs
        """)
        
        result = session.execute(query).fetchone()
        
        if result and result[0]:
            now = get_local_now()
            
            diff = now - result[0]
            total_seconds = abs(diff.total_seconds())
            
            if total_seconds < 60:
                return "Hace unos segundos"
            elif total_seconds < 3600:
                minutes = int(total_seconds // 60)
                return f"Hace {minutes} min"
            elif total_seconds < 86400:
                hours = int(total_seconds // 3600)
                return f"Hace {hours} hora{'s' if hours > 1 else ''}"
            else:
                days = int(total_seconds // 86400)
                return f"Hace {days} día{'s' if days > 1 else ''}"
        else:
            return "N/A"
            
    except Exception as e:
        print(f"Error getting last execution time: {e}")
        return "N/A"

def get_recent_errors_count(session):
    """Obtener conteo de errores usando SQL directo"""
    try:
        from sqlalchemy import text
        from datetime import timedelta
        
        yesterday = get_local_now() - timedelta(hours=24)
        
        # Contar errores en las últimas 24h
        query = text("""
            SELECT COUNT(*) FROM scraper_logs 
            WHERE timestamp >= :yesterday 
            AND log_level = 'ERROR'
        """)
        result = session.execute(query, {"yesterday": yesterday}).fetchone()
        count = int(result[0]) if result else 0
        
        return count
    except Exception as e:
        print(f"Error getting recent errors count: {e}")
        import traceback
        traceback.print_exc()
        return 0

def get_system_alerts(session):
    """Obtener alertas del sistema desde scraper_logs - último log de cada categoría"""
    from sqlalchemy import text
    
    try:
        # Usar SQL directo para evitar problemas con el enum
        sql_query = text("""
            WITH latest_logs AS (
                SELECT DISTINCT ON (LOWER(log_level))
                    LOWER(log_level) as log_level,
                    city_code,
                    offer_type,
                    message,
                    timestamp,
                    ROW_NUMBER() OVER (PARTITION BY LOWER(log_level) ORDER BY timestamp DESC) as rn
                FROM scraper_logs
                WHERE log_level IS NOT NULL
            )
            SELECT 
                ll.log_level,
                COALESCE(c.name, ll.city_code, 'Sistema') as city_name,
                ll.offer_type,
                ll.message,
                ll.timestamp
            FROM latest_logs ll
            LEFT JOIN city c ON ll.city_code = c.website_name
            WHERE ll.rn = 1
            ORDER BY ll.timestamp DESC
            LIMIT 10
        """)
        
        result = session.exec(sql_query)
        rows = result.fetchall()
        
        alerts = []
        for row in rows:
            level, city_name, offer_type, message, timestamp = row
            
            # Mapear niveles a los esperados por el frontend
            level_map = {
                'error': 'critical',
                'warning': 'warning',
                'info': 'info',
                'success': 'info',
                'debug': 'info'
            }
            
            mapped_level = level_map.get(level.lower() if level else 'info', 'info')
            
            # Formatear offer_type
            offer_display = ""
            if offer_type:
                offer_display = "Venta" if offer_type.lower() == "sell" else "Renta" if offer_type.lower() == "rent" else offer_type
                offer_display = f" [{offer_display}]"
            
            # Concatenar ciudad con tipo de oferta
            city_display = f"{city_name}{offer_display}"
            
            alerts.append({
                "level": mapped_level,
                "city": city_display,
                "message": message or f"Log de tipo {level}",
                "timestamp": timestamp.isoformat() if timestamp else get_local_now().isoformat()
            })
        
        # Si no hay alertas, agregar una informativa
        if not alerts:
            alerts.append({
                "level": "info",
                "city": "Sistema",
                "message": "No hay logs disponibles en el sistema",
                "timestamp": get_local_now().isoformat()
            })
        
        return alerts
        
    except Exception as e:
        # Retornar alerta de error si algo falla
        return [{
            "level": "warning",
            "city": "Sistema",
            "message": f"Error obteniendo alertas: {str(e)}",
            "timestamp": get_local_now().isoformat()
        }]

@app.get("/api/dashboard")
async def get_dashboard():
    """Dashboard con datos REALES de BD que funciona"""
    
    # Importar solo cuando necesitemos para evitar problemas de inicialización
    from sqlmodel import Session, select, func
    from config.db_connection import engine
    from models.city import City
    from models.scraper_log import ScraperLog
    from models.property import Property
    
    try:
        with Session(engine) as session:
            cities = session.exec(select(City).order_by(City.name)).all()
            
            # Obtener estadísticas globales reales
            total_properties_global, today_properties_global = get_property_stats(session)
            avg_speed = get_avg_speed(session)
            
            # Calcular propiedades actualizadas hoy
            from models.property import Property
            today = get_local_now().date()
            updated_today_query = select(func.count(Property.fr_property_id)).where(Property.last_update == today)
            properties_updated_today = session.exec(updated_today_query).first() or 0
            
            city_data = []
            for city in cities:
                # Calcular páginas procesadas basándose en propiedades (25 propiedades por página)
                sell_pages_processed = city.current_sell_offset // 25 if city.current_sell_offset > 0 else 0
                rent_pages_processed = city.current_rent_offset // 25 if city.current_rent_offset > 0 else 0
                
                # Limitar páginas procesadas al límite máximo
                sell_pages_processed = min(sell_pages_processed, city.sell_pages_limit)
                rent_pages_processed = min(rent_pages_processed, city.rent_pages_limit)
                
                # Calcular progreso basado en páginas procesadas vs páginas totales (máximo 100%)
                sell_progress = min((sell_pages_processed / city.sell_pages_limit * 100), 100.0) if city.sell_pages_limit > 0 else 0
                rent_progress = min((rent_pages_processed / city.rent_pages_limit * 100), 100.0) if city.rent_pages_limit > 0 else 0
                
                # Obtener estadísticas reales por ciudad
                total_properties_city, today_properties_city = get_property_stats(session, city.id)
                
                # Calcular hours_inactive reales
                hours_inactive = 0.0
                if city.last_updated:
                    delta = get_local_now().date() - city.last_updated
                    hours_inactive = round(delta.total_seconds() / 3600, 1)
                
                city_data.append({
                    "id": city.id,
                    "name": city.name,
                    "website_name": city.website_name,
                    "sell_progress": round(sell_progress, 1),
                    "rent_progress": round(rent_progress, 1),
                    "sell_pages": f"{sell_pages_processed}/{city.sell_pages_limit}",
                    "rent_pages": f"{rent_pages_processed}/{city.rent_pages_limit}",
                    "status": get_city_status(city),
                    "last_update": city.last_updated.isoformat() if city.last_updated else get_local_now().isoformat(),
                    "hours_inactive": hours_inactive,
                    "properties_today": today_properties_city,
                    "properties_total": total_properties_city
                })
            
            return {
                "status": "success",
                "timestamp": get_local_now().isoformat(),
                "data": {
                    "summary": {
                        "total_cities": len(cities),
                        "active_cities": len([c for c in cities if not c.updated]),
                        "completed_cities": len([c for c in cities if c.updated]),
                        "properties_today": today_properties_global,
                        "properties_updated_today": (properties_updated_today - today_properties_global),
                        "properties_total": total_properties_global,
                        "avg_speed_ms": avg_speed,
                        "last_execution_time": get_last_execution_time(session),
                        "recent_errors_count": get_recent_errors_count(session)
                    },
                    "cities": city_data,
                    "next_executions": get_next_executions(session),
                    "alerts": get_system_alerts(session),
                    "recent_logs": get_recent_logs(session)
                }
            }
    except Exception as e:
        return {"status": "error", "detail": f"Error: {str(e)}"}

@app.get("/api/summary")
async def get_summary_with_changes():
    """Obtener resumen con cambios porcentuales"""
    from sqlmodel import Session, select, func
    from config.db_connection import engine
    from models.city import City
    from models.property import Property
    from datetime import timedelta
    
    try:
        with Session(engine) as session:
            # Obtener estadísticas de hoy
            today = get_local_now().date()
            yesterday = today - timedelta(days=1)
            
            # Propiedades de hoy vs ayer
            today_properties_query = select(func.count(Property.fr_property_id)).where(Property.creation_date == today)
            yesterday_properties_query = select(func.count(Property.fr_property_id)).where(Property.creation_date == yesterday)
            
            today_count = session.exec(today_properties_query).first() or 0
            yesterday_count = session.exec(yesterday_properties_query).first() or 0
            
            # Calcular cambio porcentual de propiedades
            properties_change = 0
            if yesterday_count > 0:
                properties_change = round(((today_count - yesterday_count) / yesterday_count) * 100, 1)
            
            # Ciudades activas vs ayer
            cities_query = select(City)
            cities = session.exec(cities_query).all()
            
            total_cities = len(cities)
            active_cities = len([c for c in cities if not c.updated])
            
            # Para ciudades, comparar con las de ayer (simplificado)
            cities_change = 0  # Podríamos implementar historial de ciudades activas
            
            # Total de propiedades vs la semana pasada
            week_ago = today - timedelta(days=7)
            total_properties_query = select(func.count(Property.fr_property_id))
            week_ago_query = select(func.count(Property.fr_property_id)).where(Property.creation_date <= week_ago)
            
            total_properties = session.exec(total_properties_query).first() or 0
            week_ago_total = session.exec(week_ago_query).first() or 0
            
            # Calcular cambio total semanal
            total_change = 0
            if week_ago_total > 0:
                total_change = round(((total_properties - week_ago_total) / week_ago_total) * 100, 1)
            
            return {
                "status": "success",
                "data": {
                    "total_cities": total_cities,
                    "active_cities": active_cities,
                    "completed_cities": total_cities - active_cities,
                    "properties_today": today_count,
                    "properties_total": total_properties,
                    "avg_speed_ms": get_avg_speed(session),
                    "last_execution_time": get_last_execution_time(session),
                    "recent_errors_count": get_recent_errors_count(session),
                    "changes": {
                        "properties_today_change": properties_change,
                        "cities_change": cities_change,
                        "total_change": total_change
                    }
                }
            }
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.get("/api/properties")
async def get_properties(
    page: int = 1,
    limit: int = 50,
    city_ids: Optional[List[int]] = Query(None),
    offer_type: str = None,
    min_price: float = None,
    max_price: float = None,
    min_area: float = None,
    max_area: float = None,
    rooms: Optional[List[str]] = Query(None),
    baths: Optional[List[str]] = Query(None),
    garages: Optional[List[str]] = Query(None),
    stratums: Optional[List[str]] = Query(None),
    antiquity_categories: Optional[List[int]] = Query(None),
    antiquity_filter: str = None,
    property_type: Optional[List[str]] = Query(None),
    min_sale_price: float = None,
    max_sale_price: float = None,
    min_rent_price: float = None,
    max_rent_price: float = None,
    updated_date_from: str = None,
    updated_date_to: str = None,
    # Parámetros de ubicación para filtrar por distancia
    search_address: str = None,
    latitude: float = None,
    longitude: float = None,
    radius: int = None
):
    """Obtener propiedades con filtros y paginación"""
    from sqlmodel import Session, select
    from config.db_connection import engine
    from models.property import Property
    from models.city import City
    from sqlalchemy import and_, or_
    
    
    try:
        with Session(engine) as session:
            # Query base con left join a city para incluir propiedades sin city_id
            query = select(Property, City).outerjoin(City, Property.city_id == City.id)
            
            # Aplicar filtros
            filters = []
            
            if city_ids:
                filters.append(Property.city_id.in_(city_ids))
            
            if offer_type:
                filters.append(Property.offer == offer_type)
                
            if min_price is not None:
                filters.append(Property.price >= min_price)
                
            if max_price is not None:
                filters.append(Property.price <= max_price)
                
            # Filtros de precio específicos por tipo de oferta - funcionan independientemente
            sale_price_conditions = []
            rent_price_conditions = []
            
            if min_sale_price is not None:
                sale_price_conditions.append(Property.price >= min_sale_price)
            if max_sale_price is not None:
                sale_price_conditions.append(Property.price <= max_sale_price)
                
            if min_rent_price is not None:
                rent_price_conditions.append(Property.price >= min_rent_price)
            if max_rent_price is not None:
                rent_price_conditions.append(Property.price <= max_rent_price)
            
            # Aplicar filtros independientes por tipo de oferta
            price_type_conditions = []
            if sale_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'sell',
                    *sale_price_conditions
                ))
            if rent_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'rent', 
                    *rent_price_conditions
                ))
            
            if price_type_conditions:
                filters.append(or_(*price_type_conditions))
                
            if min_area is not None:
                filters.append(Property.area >= min_area)
                
            if max_area is not None:
                filters.append(Property.area <= max_area)
                
            if rooms is not None and len(rooms) > 0:
                rooms_conditions = []
                for room in rooms:
                    if room == "unspecified":  # "Sin especificar" case
                        rooms_conditions.append(or_(
                            Property.rooms == None,
                            Property.rooms == '',
                            Property.rooms == 'N/A',
                            Property.rooms == 'Sin especificar'
                        ))
                    elif room.endswith('+'):  # Casos como "5+"
                        min_value = int(room[:-1])  # Extraer el número antes del "+"
                        # Filtrar solo valores numéricos y luego comparar
                        from sqlalchemy import cast, Integer, and_
                        rooms_conditions.append(and_(
                            Property.rooms.regexp_match('^[0-9]+$'),  # Solo valores numéricos
                            cast(Property.rooms, Integer) >= min_value
                        ))
                    else:
                        rooms_conditions.append(Property.rooms == room)
                if rooms_conditions:
                    filters.append(or_(*rooms_conditions))
            
            if baths is not None and len(baths) > 0:
                baths_conditions = []
                for bath in baths:
                    if bath == "unspecified":  # "Sin especificar" case
                        baths_conditions.append(or_(
                            Property.baths == None,
                            Property.baths == '',
                            Property.baths == 'N/A',
                            Property.baths == 'Sin especificar'
                        ))
                    elif bath.endswith('+'):  # Casos como "4+"
                        min_value = int(bath[:-1])  # Extraer el número antes del "+"
                        # Filtrar solo valores numéricos y luego comparar
                        from sqlalchemy import cast, Integer, and_
                        baths_conditions.append(and_(
                            Property.baths.regexp_match('^[0-9]+$'),  # Solo valores numéricos
                            cast(Property.baths, Integer) >= min_value
                        ))
                    else:
                        baths_conditions.append(Property.baths == bath)
                if baths_conditions:
                    filters.append(or_(*baths_conditions))
            
            if garages is not None and len(garages) > 0:
                garages_conditions = []
                for garage in garages:
                    if garage == "unspecified":  # "Sin especificar" case
                        garages_conditions.append(or_(
                            Property.garages == None,
                            Property.garages == '',
                            Property.garages == 'N/A',
                            Property.garages == 'Sin especificar'
                        ))
                    elif garage.endswith('+'):  # Casos como "3+"
                        min_value = int(garage[:-1])  # Extraer el número antes del "+"
                        # Filtrar solo valores numéricos y luego comparar
                        from sqlalchemy import cast, Integer, and_
                        garages_conditions.append(and_(
                            Property.garages.regexp_match('^[0-9]+$'),  # Solo valores numéricos
                            cast(Property.garages, Integer) >= min_value
                        ))
                    else:
                        garages_conditions.append(Property.garages == garage)
                if garages_conditions:
                    filters.append(or_(*garages_conditions))
            
            if stratums is not None and len(stratums) > 0:
                stratum_conditions = []
                for stratum in stratums:
                    if stratum == "unspecified":  # "Sin especificar" case
                        stratum_conditions.append(or_(
                            Property.stratum == None,
                            Property.stratum == '',
                            Property.stratum == 'Sin especificar'
                        ))
                    else:
                        # Convertir número a formato "Estrato X"
                        stratum_str = f"Estrato {stratum}"
                        stratum_conditions.append(Property.stratum == stratum_str)
                if stratum_conditions:
                    filters.append(or_(*stratum_conditions))
            
            # Combinar filtros de antigüedad (categories y unspecified)
            if (antiquity_categories is not None and len(antiquity_categories) > 0) or antiquity_filter == 'unspecified':
                # Mapear rangos a todos los valores posibles en la BD
                from sqlalchemy import or_, cast, Integer
                antiquity_conditions = []
                
                # Procesar categories si existen
                if antiquity_categories is not None and len(antiquity_categories) > 0:
                    for antiquity_category in antiquity_categories:
                        if antiquity_category == 1:
                            # Menos de 1 año - solo valores categóricos
                            antiquity_conditions.extend([
                                Property.antiquity == 'LESS_THAN_1_YEAR',
                                Property.antiquity == 'Menos de 1 año',
                                Property.antiquity == '1'
                            ])
                        elif antiquity_category == 2:
                            # 1 a 8 años
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_1_TO_8_YEARS',
                                Property.antiquity == '1 a 8 años',
                                Property.antiquity == '2'
                            ])
                        elif antiquity_category == 3:
                            # 9 a 15 años - solo valores categóricos
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_9_TO_15_YEARS',
                                Property.antiquity == '9 a 15 años',
                                Property.antiquity == '3'
                            ])
                        elif antiquity_category == 4:
                            # 16 a 30 años - solo valores categóricos
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_16_TO_30_YEARS',
                                Property.antiquity == '16 a 30 años',
                                Property.antiquity == '4'
                            ])
                        elif antiquity_category == 5:
                            # Más de 30 años - solo valores categóricos
                            antiquity_conditions.extend([
                                Property.antiquity == 'MORE_THAN_30_YEARS',
                                Property.antiquity == 'Más de 30 años',
                                Property.antiquity == '5'
                            ])
                
                # Agregar condición unspecified si está presente
                if antiquity_filter == 'unspecified':
                    antiquity_conditions.extend([
                        Property.antiquity == 'UNDEFINED',
                        Property.antiquity == 'Sin especificar',
                        Property.antiquity == None,
                        Property.antiquity == ''
                    ])
                
                # Aplicar filtro combinado si hay condiciones
                if antiquity_conditions:
                    filters.append(or_(*antiquity_conditions))
            
            if property_type and len(property_type) > 0:
                # Manejar múltiples tipos de propiedad con búsqueda más específica
                type_conditions = []
                
                for prop_type in property_type:
                    property_type_lower = prop_type.lower()
                    if property_type_lower == "apartamento":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% apartamento %"),
                                Property.title.ilike("apartamento %"),
                                Property.title.ilike("% apartamento"),
                                Property.title.ilike("% apto %"),
                                Property.title.ilike("apto %"),
                                Property.title.ilike("% apto"),
                                Property.title.ilike("apartamento"),
                                Property.title.ilike("apto")
                            ),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "casa":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% casa %"),
                                Property.title.ilike("casa %"),
                                Property.title.ilike("% casa"),
                                Property.title.ilike("casa")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%apto%"),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "oficina":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% oficina %"),
                                Property.title.ilike("oficina %"),
                                Property.title.ilike("% oficina"),
                                Property.title.ilike("oficina")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%casa%"),
                            ~Property.title.ilike("%bodega%")
                        ))
                    elif property_type_lower == "local":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% local %"),
                                Property.title.ilike("local %"),
                                Property.title.ilike("% local"),
                                Property.title.ilike("local")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%casa%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "bodega":
                        type_conditions.append(or_(
                            Property.title.ilike("% bodega %"),
                            Property.title.ilike("bodega %"),
                            Property.title.ilike("% bodega"),
                            Property.title.ilike("bodega")
                        ))
                    elif property_type_lower == "lote":
                        type_conditions.append(or_(
                            Property.title.ilike("% lote %"),
                            Property.title.ilike("lote %"),
                            Property.title.ilike("% lote"),
                            Property.title.ilike("lote")
                        ))
                    elif property_type_lower == "finca":
                        type_conditions.append(or_(
                            Property.title.ilike("% finca %"),
                            Property.title.ilike("finca %"),
                            Property.title.ilike("% finca"),
                            Property.title.ilike("finca")
                        ))
                    else:
                        # Para otros tipos, usar búsqueda simple
                        type_conditions.append(Property.title.ilike(f"%{prop_type}%"))
                
                if type_conditions:
                    filters.append(or_(*type_conditions))
            
            if updated_date_from:
                from datetime import datetime
                try:
                    date_from = datetime.strptime(updated_date_from, "%Y-%m-%d").date()
                    filters.append(Property.last_update >= date_from)
                except:
                    pass
            
            if updated_date_to:
                from datetime import datetime
                try:
                    # Incluir todo el día hasta las 23:59:59
                    date_to = datetime.strptime(updated_date_to, "%Y-%m-%d").date()
                    filters.append(Property.last_update <= date_to)
                except:
                    pass
            
            # SI HAY FILTRO POR DISTANCIA, primero geocodificar la dirección
            if search_address and radius is not None:
                
                # Geocodificar la dirección usando Google Maps API directamente
                import requests
                import os
                
                try:
                    api_key = os.getenv('GOOGLE_API_KEY')
                    if not api_key:
                        print("❌ GOOGLE_API_KEY no encontrada")
                        latitude = longitude = None
                    else:
                        # Llamar directamente a la API de Google Maps
                        url = f"https://maps.googleapis.com/maps/api/geocode/json"
                        params = {
                            'address': search_address,
                            'key': api_key,
                            'region': 'co',
                            'language': 'es'
                        }
                        
                        response = requests.get(url, params=params)
                        data = response.json()
                        
                        if data['status'] == 'OK' and data['results']:
                            location = data['results'][0]['geometry']['location']
                            latitude = location['lat']
                            longitude = location['lng']
                            formatted_address = data['results'][0]['formatted_address']
                            print(f"📍 Coordenadas obtenidas: {latitude}, {longitude}")
                            print(f"📮 Dirección formateada: {formatted_address}")
                        else:
                            print(f"❌ No se pudo geocodificar: {data.get('status', 'Error')}")
                            latitude = longitude = None
                except Exception as e:
                    print(f"❌ Error en geocodificación: {e}")
                    latitude = longitude = None
            
            # Inicializar distance_map vacío
            distance_map = {}
            
            # Ahora aplicar filtro por distancia si tenemos coordenadas
            if latitude is not None and longitude is not None and radius is not None:
                print(f"🔍 Filtrado por distancia activado: lat={latitude}, lng={longitude}, radius={radius}m")
                from math import radians, cos, sin, asin, sqrt
                
                def calculate_distance(lat1, lng1, lat2, lng2):
                    """Calcula distancia entre dos puntos en metros usando Haversine"""
                    R = 6371000  # Radio de la Tierra en metros
                    lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])
                    dlat = lat2 - lat1
                    dlng = lng2 - lng1
                    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlng/2)**2
                    c = 2 * asin(sqrt(a))
                    return round(R * c)
                
                # 1. Obtener TODAS las propiedades (sin paginación)
                # Aplicar filtros a la consulta principal solo si hay filtros
                if filters:
                    query = query.where(and_(*filters))
                
                all_results = session.exec(query.order_by(Property.creation_date.desc())).all()
                print(f"📊 Total propiedades antes del filtro: {len(all_results)}")
                
                # 2. Calcular distancias y filtrar
                filtered_results = []
                for prop, city in all_results:
                    if prop.latitude and prop.longitude:
                        distance = calculate_distance(latitude, longitude, prop.latitude, prop.longitude)
                        if distance <= radius:
                            # Almacenar distancia en diccionario separado
                            distance_map[prop.fr_property_id] = distance
                            filtered_results.append((prop, city))
                
                # 3. Ordenar por distancia (más cercanas primero)
                filtered_results.sort(key=lambda x: distance_map.get(x[0].fr_property_id, float('inf')))
                
                # 4. Actualizar total_count
                total_count = len(filtered_results)
                print(f"✅ Propiedades encontradas dentro de {radius}m: {total_count}")
                
                # 5. Aplicar paginación DESPUÉS del filtrado
                offset = (page - 1) * limit
                results = filtered_results[offset:offset + limit]
                
            else:
                # Sin filtro de distancia, usar lógica normal
                # Contar total para paginación usando COUNT de SQL
                from sqlmodel import func
                count_query = select(func.count(Property.fr_property_id)).outerjoin(City, Property.city_id == City.id)
                if filters:
                    count_query = count_query.where(and_(*filters))
                
                total_count = session.exec(count_query).one()
                
                # Aplicar filtros a la consulta principal solo si hay filtros
                if filters:
                    query = query.where(and_(*filters))
                
                # Aplicar paginación
                offset = (page - 1) * limit
                query = query.offset(offset).limit(limit).order_by(Property.creation_date.desc())
                
                results = session.exec(query).all()
            
            # Formatear resultados
            properties = []
            
            for prop, city in results:
                # Construir link de FincaRaiz
                finca_raiz_link = None
                if hasattr(prop, 'fr_property_id') and prop.fr_property_id:
                    finca_raiz_link = f"https://www.fincaraiz.com.co/inmueble/{prop.fr_property_id}"
                
                # Construir link de Google Maps usando coordenadas
                maps_link = None
                if prop.latitude and prop.longitude:
                    maps_link = f"https://www.google.com/maps?q={prop.latitude},{prop.longitude}"
                
                # Procesar stratum (extraer número del string "Estrato X")
                stratum_value = None
                if prop.stratum:
                    import re
                    stratum_match = re.search(r'Estrato\s*(\d+)', prop.stratum)
                    if stratum_match:
                        stratum_value = int(stratum_match.group(1))
                
                # Procesar antiquity - formatear para mostrar el rango correcto
                antiquity_display = None
                if prop.antiquity:
                    # Mapeo de strings conocidos de FincaRaíz
                    antiquity_string_map = {
                        'LESS_THAN_1_YEAR': 'Menos de 1 año',
                        'FROM_1_TO_8_YEARS': '1 a 8 años',
                        'FROM_9_TO_15_YEARS': '9 a 15 años',
                        'FROM_16_TO_30_YEARS': '16 a 30 años',
                        'MORE_THAN_30_YEARS': 'Más de 30 años',
                        'NEW': 'Menos de 1 año',
                        'TO_BE_BUILT': 'En construcción',
                        '1 a 8 años': '1 a 8 años',  # Ya está en formato correcto
                        '9 a 15 años': '9 a 15 años',
                        '16 a 30 años': '16 a 30 años',
                        'Sin especificar': 'Sin especificar',
                        'UNDEFINED': 'Sin especificar'
                    }
                    
                    if prop.antiquity in antiquity_string_map:
                        antiquity_display = antiquity_string_map[prop.antiquity]
                    else:
                        # Intentar convertir a número y asignar rango
                        try:
                            # Limpiar el string por si tiene espacios o caracteres extra
                            clean_value = str(prop.antiquity).strip()
                            years = int(clean_value)
                            
                            if years == 0:
                                antiquity_display = 'Menos de 1 año'
                            elif years <= 8:
                                antiquity_display = '1 a 8 años'
                            elif years <= 15:
                                antiquity_display = '9 a 15 años'
                            elif years <= 30:
                                antiquity_display = '16 a 30 años'
                            else:
                                antiquity_display = 'Más de 30 años'
                        except:
                            # Si no se puede procesar, mostrar el valor original
                            antiquity_display = str(prop.antiquity)
                
                # Convertir rooms, baths y garages a números si son strings
                rooms_value = None
                if prop.rooms:
                    try:
                        rooms_value = int(prop.rooms) if isinstance(prop.rooms, str) else prop.rooms
                    except:
                        rooms_value = prop.rooms
                
                baths_value = None
                if prop.baths:
                    try:
                        baths_value = int(prop.baths) if isinstance(prop.baths, str) else prop.baths
                    except:
                        baths_value = prop.baths
                
                garages_value = None
                if prop.garages:
                    try:
                        garages_value = int(prop.garages) if isinstance(prop.garages, str) else prop.garages
                    except:
                        garages_value = prop.garages
                
                properties.append({
                    "id": prop.fr_property_id,
                    "city": city.name if city else "Sin especificar",
                    "area": prop.area,
                    "rooms": rooms_value,
                    "price": prop.price,
                    "offer_type": "Venta" if prop.offer == "sell" else "Renta",
                    "creation_date": prop.creation_date.isoformat() if prop.creation_date else None,
                    "last_update": prop.last_update.isoformat() if prop.last_update else None,
                    "title": prop.title,
                    "finca_raiz_link": finca_raiz_link,
                    "maps_link": maps_link,
                    "latitude": prop.latitude,
                    "longitude": prop.longitude,
                    "baths": baths_value,
                    "garages": garages_value,
                    "stratum": stratum_value,
                    "antiquity": antiquity_display,
                    "is_new": prop.is_new,
                    "address": getattr(prop, 'address', None),
                    "distance": distance_map.get(prop.fr_property_id, None)
                })
            
            total_pages = (total_count + limit - 1) // limit
            
            return {
                "status": "success",
                "data": {
                    "properties": properties,
                    "pagination": {
                        "page": page,
                        "limit": limit,
                        "total_count": total_count,
                        "total_pages": total_pages,
                        "has_next": page < total_pages,
                        "has_prev": page > 1
                    }
                }
            }
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.post("/api/properties/send-excel")
async def send_properties_excel(request: dict):
    """Enviar propiedades por email en formato Excel"""
    import smtplib
    import os
    import io
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    from datetime import datetime
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlmodel import Session, select
    from config.db_connection import engine
    from models.property import Property
    from models.city import City
    from sqlalchemy import and_, or_, cast, Integer
    
    try:
        # Obtener parámetros del request
        email_destinatario = request.get('email')
        filters = request.get('filters', {})
        
        print(f"📧 Solicitud de Excel para: {email_destinatario}")
        print(f"🔍 Filtros recibidos: {filters}")
        
        if not email_destinatario:
            return {"status": "error", "detail": "Email es requerido"}
        
        # Aplicar los mismos filtros que en el endpoint de properties
        with Session(engine) as session:
            query = select(Property, City).outerjoin(City, Property.city_id == City.id)
            
            # Aplicar filtros (mismo código que en get_properties)
            filter_conditions = []
            
            if filters.get('city_ids'):
                filter_conditions.append(Property.city_id.in_(filters['city_ids']))
            
            if filters.get('offer_type'):
                filter_conditions.append(Property.offer == filters['offer_type'])
                
            if filters.get('min_price') is not None:
                filter_conditions.append(Property.price >= filters['min_price'])
                
            if filters.get('max_price') is not None:
                filter_conditions.append(Property.price <= filters['max_price'])

            # Filtros de precio específicos por tipo de oferta
            sale_price_conditions = []
            rent_price_conditions = []
            
            if filters.get('min_sale_price') is not None:
                sale_price_conditions.append(Property.price >= filters['min_sale_price'])
            if filters.get('max_sale_price') is not None:
                sale_price_conditions.append(Property.price <= filters['max_sale_price'])
                
            if filters.get('min_rent_price') is not None:
                rent_price_conditions.append(Property.price >= filters['min_rent_price'])
            if filters.get('max_rent_price') is not None:
                rent_price_conditions.append(Property.price <= filters['max_rent_price'])
            
            # Aplicar filtros independientes por tipo de oferta
            price_type_conditions = []
            if sale_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'sell',
                    *sale_price_conditions
                ))
            if rent_price_conditions:
                price_type_conditions.append(and_(
                    Property.offer == 'rent', 
                    *rent_price_conditions
                ))
            
            if price_type_conditions:
                filter_conditions.append(or_(*price_type_conditions))
                
            if filters.get('min_area') is not None:
                filter_conditions.append(Property.area >= filters['min_area'])
                
            if filters.get('max_area') is not None:
                filter_conditions.append(Property.area <= filters['max_area'])
            
            # Filtro de habitaciones
            if filters.get('rooms') is not None and len(filters['rooms']) > 0:
                rooms_conditions = []
                for room in filters['rooms']:
                    if room == "unspecified":
                        rooms_conditions.append(or_(
                            Property.rooms == None,
                            Property.rooms == '',
                            Property.rooms == 'N/A',
                            Property.rooms == 'Sin especificar'
                        ))
                    elif room.endswith('+'):
                        min_value = int(room[:-1])
                        rooms_conditions.append(and_(
                            Property.rooms.regexp_match('^[0-9]+$'),
                            cast(Property.rooms, Integer) >= min_value
                        ))
                    else:
                        rooms_conditions.append(Property.rooms == room)
                if rooms_conditions:
                    filter_conditions.append(or_(*rooms_conditions))
            
            # Filtro de baños
            if filters.get('baths') is not None and len(filters['baths']) > 0:
                baths_conditions = []
                for bath in filters['baths']:
                    if bath == "unspecified":
                        baths_conditions.append(or_(
                            Property.baths == None,
                            Property.baths == '',
                            Property.baths == 'N/A',
                            Property.baths == 'Sin especificar'
                        ))
                    elif bath.endswith('+'):
                        min_value = int(bath[:-1])
                        baths_conditions.append(and_(
                            Property.baths.regexp_match('^[0-9]+$'),
                            cast(Property.baths, Integer) >= min_value
                        ))
                    else:
                        baths_conditions.append(Property.baths == bath)
                if baths_conditions:
                    filter_conditions.append(or_(*baths_conditions))
            
            # Filtro de garajes
            if filters.get('garages') is not None and len(filters['garages']) > 0:
                garages_conditions = []
                for garage in filters['garages']:
                    if garage == "unspecified":
                        garages_conditions.append(or_(
                            Property.garages == None,
                            Property.garages == '',
                            Property.garages == 'N/A',
                            Property.garages == 'Sin especificar'
                        ))
                    elif garage.endswith('+'):
                        min_value = int(garage[:-1])
                        garages_conditions.append(and_(
                            Property.garages.regexp_match('^[0-9]+$'),
                            cast(Property.garages, Integer) >= min_value
                        ))
                    else:
                        garages_conditions.append(Property.garages == garage)
                if garages_conditions:
                    filter_conditions.append(or_(*garages_conditions))
            
            # Filtro de estrato
            if filters.get('stratums') is not None and len(filters['stratums']) > 0:
                stratum_conditions = []
                for stratum in filters['stratums']:
                    if stratum == "unspecified":
                        stratum_conditions.append(or_(
                            Property.stratum == None,
                            Property.stratum == '',
                            Property.stratum == 'Sin especificar'
                        ))
                    else:
                        stratum_str = f"Estrato {stratum}"
                        stratum_conditions.append(Property.stratum == stratum_str)
                if stratum_conditions:
                    filter_conditions.append(or_(*stratum_conditions))
            
            # Filtro de antigüedad
            antiquity_categories = filters.get('antiquity_categories')
            antiquity_filter = filters.get('antiquity_filter')
            
            # Combinar filtros de antigüedad (categories y unspecified)
            if (antiquity_categories is not None and len(antiquity_categories) > 0) or antiquity_filter == 'unspecified':
                antiquity_conditions = []
                
                # Procesar categories si existen
                if antiquity_categories is not None and len(antiquity_categories) > 0:
                    # Asegurar que sean enteros
                    try:
                        antiquity_categories = [int(x) for x in antiquity_categories]
                    except:
                        pass
                        
                    for antiquity_category in antiquity_categories:
                        if antiquity_category == 1:
                            # Menos de 1 año
                            antiquity_conditions.extend([
                                Property.antiquity == 'LESS_THAN_1_YEAR',
                                Property.antiquity == 'Menos de 1 año',
                                Property.antiquity == '1'
                            ])
                        elif antiquity_category == 2:
                            # 1 a 8 años
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_1_TO_8_YEARS',
                                Property.antiquity == '1 a 8 años',
                                Property.antiquity == '2'
                            ])
                        elif antiquity_category == 3:
                            # 9 a 15 años
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_9_TO_15_YEARS',
                                Property.antiquity == '9 a 15 años',
                                Property.antiquity == '3'
                            ])
                        elif antiquity_category == 4:
                            # 16 a 30 años
                            antiquity_conditions.extend([
                                Property.antiquity == 'FROM_16_TO_30_YEARS',
                                Property.antiquity == '16 a 30 años',
                                Property.antiquity == '4'
                            ])
                        elif antiquity_category == 5:
                            # Más de 30 años
                            antiquity_conditions.extend([
                                Property.antiquity == 'MORE_THAN_30_YEARS',
                                Property.antiquity == 'Más de 30 años',
                                Property.antiquity == '5'
                            ])
                
                # Agregar condición unspecified si está presente
                if antiquity_filter == 'unspecified':
                    antiquity_conditions.extend([
                        Property.antiquity == 'UNDEFINED',
                        Property.antiquity == 'Sin especificar',
                        Property.antiquity == None,
                        Property.antiquity == ''
                    ])
                
                # Aplicar filtro combinado si hay condiciones
                if antiquity_conditions:
                    filter_conditions.append(or_(*antiquity_conditions))
            
            # Filtro de tipo de propiedad con búsqueda más específica
            property_type = filters.get('property_type')
            if property_type and len(property_type) > 0:
                # Manejar múltiples tipos de propiedad
                type_conditions = []
                
                for prop_type in property_type:
                    property_type_lower = prop_type.lower()
                    if property_type_lower == "apartamento":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% apartamento %"),
                                Property.title.ilike("apartamento %"),
                                Property.title.ilike("% apartamento"),
                                Property.title.ilike("% apto %"),
                                Property.title.ilike("apto %"),
                                Property.title.ilike("% apto"),
                                Property.title.ilike("apartamento"),
                                Property.title.ilike("apto")
                            ),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "casa":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% casa %"),
                                Property.title.ilike("casa %"),
                                Property.title.ilike("% casa"),
                                Property.title.ilike("casa")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%apto%"),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "oficina":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% oficina %"),
                                Property.title.ilike("oficina %"),
                                Property.title.ilike("% oficina"),
                                Property.title.ilike("oficina")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%casa%"),
                            ~Property.title.ilike("%bodega%")
                        ))
                    elif property_type_lower == "local":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% local %"),
                                Property.title.ilike("local %"),
                                Property.title.ilike("% local"),
                                Property.title.ilike("local")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%casa%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "bodega":
                        type_conditions.append(or_(
                            Property.title.ilike("% bodega %"),
                            Property.title.ilike("bodega %"),
                            Property.title.ilike("% bodega"),
                            Property.title.ilike("bodega")
                        ))
                    elif property_type_lower == "lote":
                        type_conditions.append(or_(
                            Property.title.ilike("% lote %"),
                            Property.title.ilike("lote %"),
                            Property.title.ilike("% lote"),
                            Property.title.ilike("lote")
                        ))
                    elif property_type_lower == "finca":
                        type_conditions.append(or_(
                            Property.title.ilike("% finca %"),
                            Property.title.ilike("finca %"),
                            Property.title.ilike("% finca"),
                            Property.title.ilike("finca")
                        ))
                    else:
                        # Para otros tipos, usar búsqueda simple
                        type_conditions.append(Property.title.ilike(f"%{prop_type}%"))
                
                if type_conditions:
                    filter_conditions.append(or_(*type_conditions))
            
            # Filtro de fechas de actualización
            if filters.get('updated_date_from'):
                from datetime import datetime
                date_from = datetime.strptime(filters.get('updated_date_from'), '%Y-%m-%d').date()
                filter_conditions.append(Property.last_update >= date_from)
            
            if filters.get('updated_date_to'):
                from datetime import datetime
                date_to = datetime.strptime(filters.get('updated_date_to'), '%Y-%m-%d').date()
                filter_conditions.append(Property.last_update <= date_to)
            
            if filter_conditions:
                query = query.where(and_(*filter_conditions))
            
            # Obtener propiedades
            results = session.exec(query.order_by(Property.creation_date.desc())).all()
            print(f"📊 Propiedades encontradas antes de filtrar por distancia: {len(results)}")
            
            # Procesar filtro de distancia si está presente
            distance_map = {}
            search_lat = filters.get('latitude')
            search_lng = filters.get('longitude')
            radius = filters.get('radius')
            
            if search_lat is not None and search_lng is not None and radius is not None:
                # Asegurar tipos numéricos
                try:
                    search_lat = float(search_lat)
                    search_lng = float(search_lng)
                    radius = float(radius) # Permitir float para radius también
                except (ValueError, TypeError) as e:
                    print(f"❌ Error convirtiendo coordenadas o radio a números: {e}")
                
                print(f"🔍 Procesando filtro de distancia para Excel: lat={search_lat}, lng={search_lng}, radius={radius}m")
                from math import radians, cos, sin, asin, sqrt
                
                def calculate_distance(lat1, lng1, lat2, lng2):
                    """Calcula distancia entre dos puntos en metros usando Haversine"""
                    try:
                        R = 6371000  # Radio de la Tierra en metros
                        lat1, lng1, lat2, lng2 = map(radians, [float(lat1), float(lng1), float(lat2), float(lng2)])
                        dlat = lat2 - lat1
                        dlng = lng2 - lng1
                        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlng/2)**2
                        c = 2 * asin(sqrt(a))
                        return round(R * c)
                    except Exception as e:
                        print(f"Error calculando distancia: {e}")
                        return float('inf')
                
                # Filtrar y calcular distancias - SOLO incluir propiedades dentro del radio
                filtered_results = []
                debug_count = 0
                
                for prop, city in results:
                    if prop.latitude and prop.longitude:
                        distance = calculate_distance(search_lat, search_lng, prop.latitude, prop.longitude)
                        if distance <= radius:
                            distance_map[prop.fr_property_id] = distance
                            filtered_results.append((prop, city))
                        elif debug_count < 5:
                            print(f"⚠️ Propiedad {prop.fr_property_id} fuera de rango: {distance}m > {radius}m (Lat: {prop.latitude}, Lng: {prop.longitude})")
                            debug_count += 1
                    elif debug_count < 5:
                         # Solo loguear si esperábamos coordenadas
                         # print(f"⚠️ Propiedad {prop.fr_property_id} sin coordenadas")
                         debug_count += 1
                
                # Ordenar por distancia (más cercanas primero)
                filtered_results.sort(key=lambda x: distance_map.get(x[0].fr_property_id, float('inf')))
                results = filtered_results
                print(f"✅ Propiedades filtradas por distancia: {len(results)}")
            
            # Función para formatear antigüedad
            def format_antiquity(antiquity_value):
                """Convierte valores de antigüedad a formato español estándar"""
                if not antiquity_value:
                    return "Sin especificar"
                
                # Diccionario de mapeo de strings en inglés
                english_to_spanish = {
                    'LESS_THAN_1_YEAR': '1',
                    'FROM_1_TO_8_YEARS': '2', 
                    'FROM_9_TO_15_YEARS': '3',
                    'FROM_16_TO_30_YEARS': '4',
                    'MORE_THAN_30_YEARS': '5',
                    'UNDEFINED': 'Sin especificar'
                }
                
                # Diccionario de mapeo de IDs numéricos
                id_to_spanish = {
                    1: "Menos de 1 año",
                    2: "1 a 8 años", 
                    3: "9 a 15 años",
                    4: "16 a 30 años",
                    5: "Más de 30 años"
                }
                
                # Convertir a string para procesamiento
                antiquity_str = str(antiquity_value).strip()
                
                # Si es un string en inglés, convertir a ID
                if antiquity_str in english_to_spanish:
                    mapped_value = english_to_spanish[antiquity_str]
                    if mapped_value == 'Sin especificar':
                        return mapped_value
                    antiquity_str = mapped_value
                
                # Si es un número o string numérico, convertir a español
                try:
                    antiquity_id = int(antiquity_str)
                    return id_to_spanish.get(antiquity_id, "Sin especificar")
                except ValueError:
                    # Si no es numérico y no está en el mapeo, verificar si contiene palabras clave
                    antiquity_lower = antiquity_str.lower()
                    if any(word in antiquity_lower for word in ['sin especificar', 'undefined', 'n/a', 'none']):
                        return "Sin especificar"
                    # Si contiene texto español ya formateado, devolverlo tal como está
                    return antiquity_str

            # Crear archivo Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Propiedades"
            
            # Encabezados con estilo
            headers = [
                'ID', 'Título', 'Ciudad', 'Tipo', 'Precio (COP)', 
                'Área (m²)', 'Habitaciones', 'Baños', 'Garajes', 
                'Estrato', 'Antigüedad', 'Distancia (m)', 'Fecha Creación', 
                'Última Actualización', 'FincaRaiz', 'Google Maps'
            ]
            
            # Aplicar estilos a los encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Verificar límite razonable para Excel (optimización)
            if len(results) > 50000:
                return {"status": "error", "detail": f"Demasiadas propiedades ({len(results):,}). El límite para Excel es 50,000 propiedades. Aplica más filtros."}
            
            print(f"📊 Generando Excel con {len(results):,} propiedades...")
            
            # Agregar datos en lotes para mejor rendimiento
            batch_size = 1000
            total_batches = (len(results) + batch_size - 1) // batch_size
            
            for batch_num in range(total_batches):
                start_idx = batch_num * batch_size
                end_idx = min(start_idx + batch_size, len(results))
                batch = results[start_idx:end_idx]
                
                print(f"📝 Procesando lote {batch_num + 1}/{total_batches} ({len(batch)} propiedades)")
                
                for i, (prop, city) in enumerate(batch):
                    row_idx = start_idx + i + 2  # +2 porque empezamos en fila 2 (después del header)
                    ws.cell(row=row_idx, column=1, value=prop.fr_property_id)
                    ws.cell(row=row_idx, column=2, value=prop.title)
                    ws.cell(row=row_idx, column=3, value=city.name if city else "Sin especificar")
                    ws.cell(row=row_idx, column=4, value="Venta" if prop.offer == "sell" else "Renta")
                    ws.cell(row=row_idx, column=5, value=prop.price)
                    ws.cell(row=row_idx, column=6, value=prop.area)
                    ws.cell(row=row_idx, column=7, value=prop.rooms)
                    ws.cell(row=row_idx, column=8, value=prop.baths)
                    ws.cell(row=row_idx, column=9, value=prop.garages)
                    ws.cell(row=row_idx, column=10, value=prop.stratum)
                    ws.cell(row=row_idx, column=11, value=format_antiquity(prop.antiquity))
                    
                    # Agregar distancia (columna 12)
                    distance = distance_map.get(prop.fr_property_id, None)
                    if distance is not None:
                        ws.cell(row=row_idx, column=12, value=f"{distance:,} m")
                    else:
                        ws.cell(row=row_idx, column=12, value="")
                    
                    ws.cell(row=row_idx, column=13, value=prop.creation_date.isoformat() if prop.creation_date else "")
                    ws.cell(row=row_idx, column=14, value=prop.last_update.isoformat() if prop.last_update else "")
                    
                    # Agregar hipervínculo de FincaRaiz
                    if prop.fr_property_id:
                        cell = ws.cell(row=row_idx, column=15)
                        cell.hyperlink = f"https://www.fincaraiz.com.co/inmueble/{prop.fr_property_id}"
                        cell.value = "Ver en FincaRaiz"
                        cell.style = "Hyperlink"
                    else:
                        ws.cell(row=row_idx, column=15, value="")
                    
                    # Agregar hipervínculo de Google Maps
                    if prop.latitude and prop.longitude:
                        cell = ws.cell(row=row_idx, column=16)
                        cell.hyperlink = f"https://www.google.com/maps?q={prop.latitude},{prop.longitude}"
                        cell.value = "Ver en Maps"
                        cell.style = "Hyperlink"
                    else:
                        ws.cell(row=row_idx, column=16, value="")
            
            # Ajustar ancho de columnas con valores predefinidos (más rápido)
            column_widths = {
                'A': 12,  # ID
                'B': 40,  # Título  
                'C': 15,  # Ciudad
                'D': 10,  # Tipo
                'E': 15,  # Precio
                'F': 12,  # Área
                'G': 12,  # Habitaciones
                'H': 8,   # Baños
                'I': 8,   # Garajes
                'J': 8,   # Estrato
                'K': 15,  # Antigüedad
                'L': 15,  # Distancia
                'M': 12,  # Fecha Creación
                'N': 12,  # Última Actualización
                'O': 15,  # FincaRaiz
                'P': 15   # Google Maps
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Guardar Excel en memoria
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Enviar email
            print(f"📧 Enviando email a {email_destinatario}...")
            
            # Configuración del servidor SMTP
            smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
            smtp_port = int(os.getenv("SMTP_PORT", "587"))
            smtp_user = os.getenv("SMTP_USER")
            smtp_password = os.getenv("SMTP_PASSWORD")
            from_email = os.getenv("FROM_EMAIL", smtp_user)
            
            if not smtp_user or not smtp_password:
                return {"status": "error", "detail": "Credenciales SMTP no configuradas"}
            
            # Crear mensaje
            msg = MIMEMultipart()
            msg['From'] = from_email
            msg['To'] = email_destinatario
            msg['Cc'] = from_email
            msg['Subject'] = f"Dashboard Scraper - Propiedades Exportadas"
            
            # Cuerpo del correo
            # Agregar información de búsqueda por dirección si aplica
            address_info = ""
            if search_lat is not None and search_lng is not None and filters.get('search_address'):
                address_info = f"\n• Propiedad consultada: {filters.get('search_address')}"
                if radius:
                    address_info += f"\n• Radio de búsqueda: {radius:,} metros"
            
            body = f"""Dashboard Scraper - Propiedades

Adjunto encontrarás el archivo Excel con las propiedades solicitadas.

Resumen:
• Total de propiedades exportadas: {len(results)}
• Fecha de exportación: {get_local_now().strftime("%d/%m/%Y %H:%M")}{address_info}

El archivo Excel contiene información detallada de cada propiedad incluyendo:
• Información básica (título, ciudad, tipo)
• Detalles de precio y características  
• Enlaces directos a FincaRaiz
• Coordenadas geográficas

Si tienes alguna pregunta, no dudes en contactarnos.

Saludos cordiales,
Equipo de Avalúos"""
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Adjuntar Excel
            filename = f"propiedades_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.set_payload(excel_file.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(attachment)
            
            # Enviar correo
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
            
            print(f"✅ Email enviado exitosamente a {email_destinatario}")
            
            return {
                "status": "success",
                "message": f"Excel enviado exitosamente a {email_destinatario}",
                "properties_count": len(results)
            }
            
    except Exception as e:
        print(f"❌ Error enviando email: {e}")
        return {"status": "error", "detail": str(e)}

@app.get("/api/zone-statistics")
async def get_zone_statistics_v2(
    city_id: int = None,
    updated_date_from: str = None,
    updated_date_to: str = None
):
    """Obtener estadísticas de precio por zona para el mapa con valorización"""
    from sqlmodel import Session, select, func
    from config.db_connection import engine  
    from models.property import Property
    from datetime import datetime
    from sqlalchemy import text
    
    try:
        with Session(engine) as session:
            # Query optimizada para obtener solo coordenadas de zonas
            query_sql = """
            SELECT 
                p.location_main,
                COUNT(DISTINCT p.fr_property_id) as property_count,
                MIN(p.latitude) as min_lat,
                MAX(p.latitude) as max_lat,
                MIN(p.longitude) as min_lng,
                MAX(p.longitude) as max_lng,
                AVG(p.latitude) as center_lat,
                AVG(p.longitude) as center_lng
            FROM property p
            WHERE p.location_main IS NOT NULL 
                AND p.area > 0 
                AND p.price > 0
                AND p.latitude IS NOT NULL 
                AND p.longitude IS NOT NULL
                {city_filter}
            GROUP BY p.location_main
            HAVING COUNT(DISTINCT p.fr_property_id) > 3
            ORDER BY property_count DESC;
            """
            
            # Aplicar filtros dinámicos
            city_filter = ""
            if city_id:
                city_filter = f"AND p.city_id = {city_id}"
            
            final_query = query_sql.format(city_filter=city_filter)
            
            results = session.exec(text(final_query)).all()
            
            zone_stats = []
            for row in results:
                # Solo devolver información básica de zonas con sus límites
                zone_stats.append({
                    'id': row[0].lower().replace(' ', '_').replace('/', '_') if row[0] else 'unknown',
                    'name': row[0] if row[0] else 'Zona desconocida',
                    'property_count': int(row[1]) if row[1] else 0,
                    'bounds': {
                        'min_lat': float(row[2]) if row[2] else 0,
                        'max_lat': float(row[3]) if row[3] else 0,
                        'min_lng': float(row[4]) if row[4] else 0,
                        'max_lng': float(row[5]) if row[5] else 0
                    },
                    'center_lat': float(row[6]) if row[6] else 0,
                    'center_lng': float(row[7]) if row[7] else 0,
                    # Los valores se calcularán al hacer clic
                    'sale_avg_price_m2': 0,
                    'sale_valorization': 0,
                    'rent_avg_price_m2': 0,
                    'rent_valorization': 0,
                    'cap_rate': 0,
                    'cap_rate_valorization': 0
                })
            
            return {
                'status': 'success',
                'data': zone_stats
            }
            
    except Exception as e:
        print(f"Error getting zone statistics: {e}")
        return {
            'status': 'error',
            'message': str(e),
            'data': []
        }

@app.get("/api/all-postal-codes")
async def get_all_postal_codes_for_city(city_id: int = None):
    """
    Obtener códigos postales de una ciudad.
    Por ahora solo devuelve los que tienen propiedades en la BD.
    """
    from sqlmodel import Session
    from config.db_connection import engine
    from sqlalchemy import text
    
    try:
        # La tabla no tiene postal_code, vamos a usar location_main como zona
        with Session(engine) as session:
            query = """
                SELECT DISTINCT 
                    p.location_main,
                    COUNT(*) as property_count,
                    AVG(p.latitude) as center_lat,
                    AVG(p.longitude) as center_lng,
                    AVG(CASE WHEN p.offer = 'sell' THEN p.price / NULLIF(p.area, 0) END) as avg_sale_price_m2,
                    AVG(CASE WHEN p.offer = 'rent' THEN p.price / NULLIF(p.area, 0) END) as avg_rent_price_m2,
                    AVG(p.price / NULLIF(p.area, 0)) as avg_price_m2
                FROM property p
                WHERE p.location_main IS NOT NULL
                    AND p.location_main != ''
                    AND p.latitude IS NOT NULL
                    AND p.longitude IS NOT NULL
                    {}
                GROUP BY p.location_main
                ORDER BY property_count DESC
            """
            
            city_filter = f"AND p.city_id = {city_id}" if city_id else ""
            result = session.exec(text(query.format(city_filter))).all()
            
            all_postal_codes = []
            for row in result:
                all_postal_codes.append({
                    'postal_code': row[0],  # Usando location_main como "código postal"
                    'has_properties': True,
                    'property_count': row[1],
                    'center_lat': float(row[2]) if row[2] else None,
                    'center_lng': float(row[3]) if row[3] else None,
                    'avg_sale_price_m2': float(row[4]) if row[4] else 0,
                    'avg_rent_price_m2': float(row[5]) if row[5] else 0,
                    'avg_price_m2': float(row[6]) if row[6] else 0
                })
        
        return {
            'status': 'success',
            'data': all_postal_codes,
            'stats': {
                'total_codes': len(all_postal_codes),
                'codes_with_properties': len(all_postal_codes),
                'codes_without_properties': 0  # Por ahora solo mostramos los que tienen propiedades
            }
        }
        
    except Exception as e:
        print(f"Error getting all postal codes: {e}")
        return {
            'status': 'error',
            'message': str(e),
            'data': []
        }

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "ok", "message": "Backend is running"}

# All postal-cache endpoints removed - using Google Maps Data-Driven Styling only
# OSM/Overture/Geoapify endpoints removed - using Google Maps only

@app.get("/api/zone-statistics-full")
async def get_zone_statistics_full(
    city_id: int = None,
    updated_date_from: str = None,
    updated_date_to: str = None
):
    """Obtener estadísticas completas de zonas para el mapa con colores"""
    from sqlmodel import Session
    from config.db_connection import engine  
    from sqlalchemy import text
    
    try:
        with Session(engine) as session:
            query_sql = """
            WITH zone_stats AS (
                SELECT 
                    p.location_main,
                    c.name as city_name,
                    COUNT(DISTINCT p.fr_property_id) as property_count,
                    -- Usar percentiles para obtener un área más representativa (excluir outliers)
                    PERCENTILE_CONT(0.2) WITHIN GROUP (ORDER BY p.latitude) as min_lat,
                    PERCENTILE_CONT(0.8) WITHIN GROUP (ORDER BY p.latitude) as max_lat,
                    PERCENTILE_CONT(0.2) WITHIN GROUP (ORDER BY p.longitude) as min_lng,
                    PERCENTILE_CONT(0.8) WITHIN GROUP (ORDER BY p.longitude) as max_lng,
                    AVG(p.latitude) as center_lat,
                    AVG(p.longitude) as center_lng,
                    -- Precios actuales
                    AVG(CASE WHEN p.offer = 'sell' THEN p.price / NULLIF(p.area, 0) END) as sale_price_m2,
                    AVG(CASE WHEN p.offer = 'rent' THEN p.price / NULLIF(p.area, 0) END) as rent_price_m2,
                    -- Precios previos para valorización
                    AVG(CASE WHEN p.offer = 'sell' AND up.previous_value > 0 THEN up.previous_value / NULLIF(p.area, 0) END) as prev_sale_m2,
                    AVG(CASE WHEN p.offer = 'rent' AND up.previous_value > 0 THEN up.previous_value / NULLIF(p.area, 0) END) as prev_rent_m2
                FROM property p
                LEFT JOIN updated_property up ON p.fr_property_id = up.property_id
                LEFT JOIN city c ON p.city_id = c.id
                WHERE p.location_main IS NOT NULL 
                    AND p.area > 0 
                    AND p.price > 0
                    AND p.latitude IS NOT NULL 
                    AND p.longitude IS NOT NULL
                    {city_filter}
                    {date_filter}
                GROUP BY p.location_main, c.name
                HAVING COUNT(DISTINCT p.fr_property_id) > 3
            )
            SELECT 
                location_main as name,
                city_name,
                property_count,
                min_lat, max_lat, min_lng, max_lng, center_lat, center_lng,
                COALESCE(sale_price_m2, 0) as sale_price_m2,
                COALESCE(rent_price_m2, 0) as rent_price_m2,
                CASE 
                    WHEN prev_sale_m2 > 0 AND sale_price_m2 > 0 THEN 
                        ((sale_price_m2 - prev_sale_m2) / prev_sale_m2 * 100)
                    ELSE 0 
                END as sale_valorization,
                CASE 
                    WHEN prev_rent_m2 > 0 AND rent_price_m2 > 0 THEN 
                        ((rent_price_m2 - prev_rent_m2) / prev_rent_m2 * 100)
                    ELSE 0 
                END as rent_valorization,
                CASE 
                    WHEN sale_price_m2 > 0 AND rent_price_m2 > 0 THEN 
                        (rent_price_m2 / sale_price_m2 / 12)
                    ELSE 0 
                END as cap_rate
            FROM zone_stats
            ORDER BY property_count DESC;
            """
            
            city_filter = f"AND p.city_id = {city_id}" if city_id else ""
            
            date_filter = ""
            if updated_date_from:
                date_filter += f" AND (up.updated_date IS NULL OR up.updated_date >= '{updated_date_from}')"
            if updated_date_to:
                date_filter += f" AND (up.updated_date IS NULL OR up.updated_date <= '{updated_date_to}')"
                
            final_query = query_sql.format(city_filter=city_filter, date_filter=date_filter)
            
            results = session.exec(text(final_query)).all()
            
            zones_data = []
            for result in results:
                import math
                
                # Asegurar que no hay NaN - ajustando índices por city_name
                sale_price_m2 = float(result[9]) if result[9] and not math.isnan(float(result[9])) else 0
                rent_price_m2 = float(result[10]) if result[10] and not math.isnan(float(result[10])) else 0
                sale_valorization = float(result[11]) if result[11] and not math.isnan(float(result[11])) else 0
                rent_valorization = float(result[12]) if result[12] and not math.isnan(float(result[12])) else 0
                cap_rate = float(result[13]) if result[13] and not math.isnan(float(result[13])) else 0
                
                zones_data.append({
                    'id': str(result[0]),
                    'name': str(result[0]),
                    'city_name': str(result[1]) if result[1] else '',
                    'property_count': int(result[2]) if result[2] else 0,
                    'min_lat': float(result[3]) if result[3] else 0,
                    'max_lat': float(result[4]) if result[4] else 0,
                    'min_lng': float(result[5]) if result[5] else 0,
                    'max_lng': float(result[6]) if result[6] else 0,
                    'center_lat': float(result[7]) if result[7] else 0,
                    'center_lng': float(result[8]) if result[8] else 0,
                    'sale_price_m2': sale_price_m2,
                    'rent_price_m2': rent_price_m2,
                    'sale_valorization': sale_valorization,
                    'rent_valorization': rent_valorization,
                    'cap_rate': cap_rate
                })
            
            return {
                'status': 'success',
                'data': zones_data
            }
                
    except Exception as e:
        print(f"Error getting zone statistics full: {e}")
        return {'status': 'error', 'message': str(e)}

@app.get("/api/zone-details")
async def get_zone_details(
    zone_name: str,
    city_id: int = None,
    updated_date_from: str = None,
    updated_date_to: str = None,
    property_type: str = None,
    # Coordenadas del bounding box de la zona
    north: float = None,
    south: float = None,
    east: float = None,
    west: float = None
):
    """Obtener detalles y estadísticas de una zona específica con comparación de períodos"""
    from sqlmodel import Session
    from config.db_connection import engine  
    from sqlalchemy import text
    from datetime import datetime, timedelta
    
    try:
        with Session(engine) as session:
            # Verificar si tenemos coordenadas
            if not all([north, south, east, west]):
                # Fallback a búsqueda por nombre
                location_filter = "p.location_main = :zone_name"
                location_params = {"zone_name": zone_name}
            else:
                # Usar coordenadas para filtrar
                location_filter = """
                    p.latitude IS NOT NULL 
                    AND p.longitude IS NOT NULL
                    AND p.latitude <= :north 
                    AND p.latitude >= :south
                    AND p.longitude <= :east 
                    AND p.longitude >= :west
                """
                location_params = {
                    "north": north,
                    "south": south,
                    "east": east,
                    "west": west
                }
            
            # Determinar si hay filtro de fecha para hacer comparación
            has_date_filter = updated_date_from or updated_date_to
            
            # PERÍODO FILTRADO (o actual si no hay filtro)
            # Usar previous_value de updated_property si existe, sino precio de property
            query_filtered = f"""
            WITH zone_data AS (
                SELECT 
                    p.fr_property_id,
                    p.offer,
                    COALESCE(up.previous_value, p.price) as price,
                    p.area,
                    p.creation_date
                FROM property p
                LEFT JOIN updated_property up ON p.fr_property_id = up.property_id
                WHERE {location_filter}
                    AND p.area > 0 
                    AND (COALESCE(up.previous_value, p.price) > 0)
                    {{city_filter}}
                    {{date_filter}}
                    {{type_filter}}
            ),
            area_stats AS (
                SELECT
                    AVG(area) as mean_area,
                    STDDEV(area) as stddev_area
                FROM zone_data
            ),
            price_stats AS (
                SELECT
                    offer,
                    AVG(price) as mean_price,
                    STDDEV(price) as stddev_price
                FROM zone_data
                GROUP BY offer
            ),
            filtered_data AS (
                SELECT zd.*
                FROM zone_data zd
                CROSS JOIN area_stats ast
                LEFT JOIN price_stats ps ON zd.offer = ps.offer
                WHERE zd.price BETWEEN (ps.mean_price - 3 * COALESCE(ps.stddev_price, 0)) 
                                   AND (ps.mean_price + 3 * COALESCE(ps.stddev_price, 0))
                  AND zd.area BETWEEN (ast.mean_area - 3 * COALESCE(ast.stddev_area, 0))
                                  AND (ast.mean_area + 3 * COALESCE(ast.stddev_area, 0))
            )
            SELECT 
                COUNT(DISTINCT fr_property_id) as total_properties,
                COUNT(DISTINCT CASE WHEN offer = 'sell' THEN fr_property_id END) as sale_count,
                COUNT(DISTINCT CASE WHEN offer = 'rent' THEN fr_property_id END) as rent_count,
                AVG(CASE WHEN offer = 'sell' THEN price / NULLIF(area, 0) END) as sale_price_m2,
                AVG(CASE WHEN offer = 'rent' THEN price / NULLIF(area, 0) END) as rent_price_m2,
                AVG(CASE WHEN offer = 'sell' THEN price END) as avg_sale_price,
                AVG(CASE WHEN offer = 'rent' THEN price END) as avg_rent_price
            FROM filtered_data;
            """
            
            city_filter = f"AND p.city_id = {city_id}" if city_id else ""
            
            # Filtro por tipo de propiedad
            type_filter = ""
            if property_type:
                property_types = [pt.strip().lower() for pt in property_type.split(',')]
                type_conditions = []
                for prop_type in property_types:
                    if prop_type == "apartamento":
                        type_conditions.append("(p.title ILIKE '%apartamento%' OR p.title ILIKE '%apto%')")
                    elif prop_type == "casa":
                        type_conditions.append("p.title ILIKE '%casa%'")
                    elif prop_type == "oficina":
                        type_conditions.append("p.title ILIKE '%oficina%'")
                    elif prop_type == "local":
                        type_conditions.append("p.title ILIKE '%local%'")
                    elif prop_type == "bodega":
                        type_conditions.append("p.title ILIKE '%bodega%'")
                    elif prop_type == "lote":
                        type_conditions.append("p.title ILIKE '%lote%'")
                    elif prop_type == "finca":
                        type_conditions.append("p.title ILIKE '%finca%'")
                if type_conditions:
                    type_filter = f"AND ({' OR '.join(type_conditions)})"
            
            # Filtro por CREATION_DATE (no last_update)
            date_filter = ""
            if updated_date_from:
                date_filter += f" AND p.creation_date >= '{updated_date_from}'"
            if updated_date_to:
                date_filter += f" AND p.creation_date <= '{updated_date_to}'"
                
            final_query_filtered = query_filtered.format(city_filter=city_filter, date_filter=date_filter, type_filter=type_filter)
            result_filtered = session.execute(text(final_query_filtered), location_params).first()
            
            # PERÍODO ACTUAL (último mes) - solo si hay filtro de fecha
            result_current = None
            if has_date_filter:
                # Calcular fecha de hace 30 días
                date_30_days_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                
                query_current = f"""
                WITH zone_data AS (
                    SELECT 
                        p.fr_property_id,
                        p.offer,
                        p.price,
                        p.area,
                        p.last_update
                    FROM property p
                    WHERE {location_filter}
                        AND p.area > 0 
                        AND p.price > 0
                        {{city_filter}}
                        {{type_filter}}
                        AND p.last_update >= :date_30_days_ago
                ),
                area_stats AS (
                    SELECT
                        AVG(area) as mean_area,
                        STDDEV(area) as stddev_area
                    FROM zone_data
                ),
                price_stats AS (
                    SELECT
                        offer,
                        AVG(price) as mean_price,
                        STDDEV(price) as stddev_price
                    FROM zone_data
                    GROUP BY offer
                ),
                filtered_data AS (
                    SELECT zd.*
                    FROM zone_data zd
                    CROSS JOIN area_stats ast
                    LEFT JOIN price_stats ps ON zd.offer = ps.offer
                    WHERE zd.price BETWEEN (ps.mean_price - 3 * COALESCE(ps.stddev_price, 0)) 
                                       AND (ps.mean_price + 3 * COALESCE(ps.stddev_price, 0))
                      AND zd.area BETWEEN (ast.mean_area - 3 * COALESCE(ast.stddev_area, 0))
                                      AND (ast.mean_area + 3 * COALESCE(ast.stddev_area, 0))
                )
                SELECT 
                    COUNT(DISTINCT fr_property_id) as total_properties,
                    COUNT(DISTINCT CASE WHEN offer = 'sell' THEN fr_property_id END) as sale_count,
                    COUNT(DISTINCT CASE WHEN offer = 'rent' THEN fr_property_id END) as rent_count,
                    AVG(CASE WHEN offer = 'sell' THEN price / NULLIF(area, 0) END) as sale_price_m2,
                    AVG(CASE WHEN offer = 'rent' THEN price / NULLIF(area, 0) END) as rent_price_m2,
                    AVG(CASE WHEN offer = 'sell' THEN price END) as avg_sale_price,
                    AVG(CASE WHEN offer = 'rent' THEN price END) as avg_rent_price
                FROM filtered_data;
                """
                
                final_query_current = query_current.format(city_filter=city_filter, type_filter=type_filter)
                current_params = {**location_params, "date_30_days_ago": date_30_days_ago}
                result_current = session.execute(text(final_query_current), current_params).first()
            
            # Procesar resultado del período filtrado
            filtered_data = {}
            if result_filtered:
                sale_count = int(result_filtered[1]) if result_filtered[1] else 0
                rent_count = int(result_filtered[2]) if result_filtered[2] else 0
                sale_price_m2 = float(result_filtered[3]) if result_filtered[3] else 0
                rent_avg_price = float(result_filtered[4]) if result_filtered[4] else 0
                avg_sale_price = float(result_filtered[5]) if result_filtered[5] else 0
                avg_rent_price = float(result_filtered[6]) if result_filtered[6] else 0
                
                # Cap rate
                cap_rate = ((avg_rent_price * 12) / avg_sale_price) if avg_sale_price > 0 and avg_rent_price > 0 else 0
                
                # Asegurar que no hay NaN
                import math
                cap_rate = 0 if math.isnan(cap_rate) or math.isinf(cap_rate) else cap_rate
                
                filtered_data = {
                    'property_count': int(result_filtered[0]) if result_filtered[0] else 0,
                    'sale_count': sale_count,
                    'rent_count': rent_count,
                    'sale_avg_price_m2': sale_price_m2,
                    'rent_avg_price_m2': rent_avg_price,
                    'cap_rate': cap_rate
                }
            
            # Procesar resultado del período actual (último mes)
            current_data = None
            if result_current and has_date_filter:
                sale_count_current = int(result_current[1]) if result_current[1] else 0
                rent_count_current = int(result_current[2]) if result_current[2] else 0
                sale_price_m2_current = float(result_current[3]) if result_current[3] else 0
                rent_avg_price_current = float(result_current[4]) if result_current[4] else 0
                avg_sale_price_current = float(result_current[5]) if result_current[5] else 0
                avg_rent_price_current = float(result_current[6]) if result_current[6] else 0
                
                # Cap rate actual
                cap_rate_current = ((avg_rent_price_current * 12) / avg_sale_price_current) if avg_sale_price_current > 0 and avg_rent_price_current > 0 else 0
                
                import math
                cap_rate_current = 0 if math.isnan(cap_rate_current) or math.isinf(cap_rate_current) else cap_rate_current
                
                property_count_current = int(result_current[0]) if result_current[0] else 0
                
                # Solo crear current_data si tiene propiedades
                if property_count_current > 0:
                    current_data = {
                        'property_count': property_count_current,
                        'sale_count': sale_count_current,
                        'rent_count': rent_count_current,
                        'sale_avg_price_m2': sale_price_m2_current,
                        'rent_avg_price_m2': rent_avg_price_current,
                        'cap_rate': cap_rate_current
                    }
                    print(f"✅ Período actual (último mes): {property_count_current} propiedades en {zone_name}")
                else:
                    print(f"⚠️ Sin propiedades en último mes para {zone_name}")
            
            # has_comparison solo True si tenemos AMBOS períodos con datos
            has_comparison = has_date_filter and current_data is not None and filtered_data.get('property_count', 0) > 0
            
            print(f"📊 Zone: {zone_name}, has_date_filter: {has_date_filter}, has_comparison: {has_comparison}")
            print(f"   Filtered period: {filtered_data.get('property_count', 0)} props")
            print(f"   Current period: {current_data.get('property_count', 0) if current_data else 0} props")
            
            return {
                'status': 'success',
                'data': {
                    'filtered_period': filtered_data if filtered_data else {
                        'property_count': 0,
                        'sale_avg_price_m2': 0,
                        'rent_avg_price_m2': 0,
                        'cap_rate': 0
                    },
                    'current_period': current_data,  # None si no hay datos
                    'has_comparison': has_comparison
                }
            }
    except Exception as e:
        print(f"Error getting zone details: {e}")
        import traceback
        traceback.print_exc()
        return {'status': 'error', 'message': str(e)}
    except Exception as e:
        print(f"Error getting zone details: {e}")
        return {'status': 'error', 'message': str(e)}

@app.get("/api/test-zones")
async def test_zone_stats():
    """Test endpoint para zona statistics"""
    try:
        from sqlmodel import Session, select, func
        from config.db_connection import engine
        from models.property import Property
        
        with Session(engine) as session:
            query = select(func.count(Property.location_main))
            result = session.exec(query).first()
            
            return {
                'status': 'success',
                'test': 'working',
                'count': result
            }
    except Exception as e:
        return {'status': 'error', 'message': str(e)}

@app.get("/api/cities/list")
async def get_cities_list():
    """Obtener lista de ciudades para filtros"""
    from sqlmodel import Session, select
    from config.db_connection import engine
    from models.city import City
    
    try:
        with Session(engine) as session:
            cities_query = select(City.id, City.name).order_by(City.name)
            cities = session.exec(cities_query).all()
            
            return {
                "status": "success",
                "data": [
                    {"id": city_id, "name": city_name}
                    for city_id, city_name in cities
                ]
            }
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.get("/api/postal-zone-statistics")
async def get_postal_zone_statistics(
    city_id: int = None,
    updated_date_from: str = None,
    updated_date_to: str = None
):
    """Obtener estadísticas de zonas agrupadas por código postal"""
    from sqlmodel import Session
    from config.db_connection import engine
    # from services.google_postal_service import GooglePostalService  # Temporarily disabled
    from sqlalchemy import text
    import json
    
    try:
        with Session(engine) as session:
            # First, get properties with coordinates
            properties_query = """
            SELECT 
                p.fr_property_id,
                p.latitude,
                p.longitude,
                p.location_main,
                c.name as city_name,
                p.city_id
            FROM property p
            LEFT JOIN city c ON p.city_id = c.id
            WHERE p.latitude IS NOT NULL 
                AND p.longitude IS NOT NULL
                AND p.area > 0 
                AND p.price > 0
                {city_filter}
            """
            
            # Apply city filter
            city_filter = ""
            if city_id:
                city_filter = f"AND p.city_id = {city_id}"
            
            final_query = properties_query.format(city_filter=city_filter)
            properties = session.exec(text(final_query)).all()
            
            print(f"🔍 Found {len(properties)} properties for postal code lookup")
            
            # Use Google Maps smart batch processing for postal codes
            coordinates = [
                {
                    'id': str(prop[0]), 
                    'lat': prop[1], 
                    'lng': prop[2]
                } for prop in properties
            ]
            
            # postal_codes_map = GooglePostalService.batch_get_postal_codes_smart(coordinates)  # Temporarily disabled
            postal_codes_map = {}  # Fallback empty map
            
            # Group properties by postal code
            postal_zones = {}
            
            for prop in properties:
                property_id = str(prop[0])
                postal_result = postal_codes_map.get(property_id)
                
                if postal_result and postal_result.postal_code:
                    postal_code = postal_result.postal_code
                    
                    if postal_code not in postal_zones:
                        postal_zones[postal_code] = {
                            'postal_code': postal_code,
                            'department': postal_result.department if postal_result else '',  # Fallback
                            'properties': [],
                            'cities': set(),
                            'neighborhoods': set()
                        }
                    
                    postal_zones[postal_code]['properties'].append({
                        'lat': prop[1],
                        'lng': prop[2],
                        'location_main': prop[3]
                    })
                    
                    if prop[4]:  # city_name
                        postal_zones[postal_code]['cities'].add(prop[4])
                    if prop[3]:  # location_main 
                        postal_zones[postal_code]['neighborhoods'].add(prop[3])
            
            # Create zone statistics for each postal code
            zone_stats = []
            
            for postal_code, data in postal_zones.items():
                properties = data['properties']
                
                # Incluir TODAS las zonas, incluso con 1 sola propiedad
                if len(properties) >= 1:  # Cambio: de 3 a 1 para incluir todas las zonas
                    lats = [p['lat'] for p in properties]
                    lngs = [p['lng'] for p in properties]
                    
                    zone_stats.append({
                        'id': f'postal_{postal_code}',
                        'name': f'{postal_code} - {data["department"]}',
                        'postal_code': postal_code,
                        'department': data['department'],
                        'cities': list(data['cities']),
                        'neighborhoods': list(data['neighborhoods']),
                        'property_count': len(properties),
                        'min_lat': min(lats),
                        'max_lat': max(lats),
                        'min_lng': min(lngs),
                        'max_lng': max(lngs),
                        'center_lat': sum(lats) / len(lats),
                        'center_lng': sum(lngs) / len(lngs)
                    })
            
            # Sort by property count
            zone_stats.sort(key=lambda x: x['property_count'], reverse=True)
            
            # Log estadísticas detalladas
            zones_by_count = {}
            for zone in zone_stats:
                count = zone['property_count']
                if count not in zones_by_count:
                    zones_by_count[count] = 0
                zones_by_count[count] += 1
            
            print(f"✅ Created {len(zone_stats)} postal code zones")
            print("📊 Zones by property count:")
            for count in sorted(zones_by_count.keys(), reverse=True)[:10]:  # Top 10
                print(f"   {count} properties: {zones_by_count[count]} zones")
            if len(zones_by_count) > 10:
                single_prop_zones = sum(zones_by_count.get(i, 0) for i in range(1, 4))
                print(f"   1-3 properties: {single_prop_zones} zones (now included!)")
            
            return {
                'status': 'success',
                'data': zone_stats,
                'total_zones': len(zone_stats),
                'total_properties': sum(z['property_count'] for z in zone_stats)
            }
            
    except Exception as e:
        print(f"Error getting postal zone statistics: {e}")
        return {
            'status': 'error',
            'message': str(e),
            'data': []
        }

@app.get("/health")
async def health():
    """Endpoint simple para healthcheck de Docker"""
    return {"status": "healthy", "timestamp": get_local_now()}

@app.get("/api/health")
async def api_health():
    """Endpoint de health con más detalles para la API"""
    return {"status": "healthy", "timestamp": get_local_now()}

# Overture Maps endpoints removed - using Google Maps Data-Driven Styling only

# Real boundaries endpoints removed - using Google Maps only

@app.get("/api/properties/by-zone")
async def get_properties_by_zone(
    boundary_type: str = Query(..., description="Tipo de límite: country, admin_level_1, admin_level_2, postal_code"),
    city_id: Optional[int] = Query(None, description="Filtrar por ciudad específica"),
    # Parámetros de bounding box para filtrar por área geográfica
    north: Optional[float] = Query(None, description="Límite norte del área"),
    south: Optional[float] = Query(None, description="Límite sur del área"),
    east: Optional[float] = Query(None, description="Límite este del área"),
    west: Optional[float] = Query(None, description="Límite oeste del área"),
    # Filtros adicionales
    property_type: Optional[str] = Query(None, description="Tipos de inmueble separados por coma (apartamento,casa)"),
    updated_date_from: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    updated_date_to: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)")
):
    """
    Obtener propiedades agrupadas por zona administrativa con estadísticas
    Retorna: lista de zonas con sus propiedades, precios promedio de venta/arriendo y cap rate
    """
    from sqlmodel import Session, select
    from config.db_connection import engine
    from models.property import Property
    from models.city import City
    from sqlalchemy import and_, or_
    
    print(f"\n🏘️ === PROPERTIES BY ZONE REQUEST ===")
    print(f"   Boundary Type: {boundary_type}")
    print(f"   City ID: {city_id}")
    print(f"   Property Type: '{property_type}'")
    print(f"   Property Type is None: {property_type is None}")
    print(f"   Property Type is empty string: {property_type == ''}")
    print(f"   Date From: {updated_date_from}")
    print(f"   Date To: {updated_date_to}")
    print(f"   Bounding Box: N={north}, S={south}, E={east}, W={west}")
    
    try:
        with Session(engine) as session:
            # Query base: obtener propiedades con coordenadas
            filters = [
                Property.latitude.isnot(None),
                Property.longitude.isnot(None)
            ]
            
            # Filtrar por ciudad si se especifica
            if city_id:
                filters.append(Property.city_id == city_id)
            
            # Filtrar por bounding box si se especifica (para optimizar la consulta)
            if north is not None and south is not None and east is not None and west is not None:
                filters.extend([
                    Property.latitude <= north,
                    Property.latitude >= south,
                    Property.longitude <= east,
                    Property.longitude >= west
                ])
            
            # Filtrar por tipos de inmueble (usando búsqueda en title)
            if property_type:
                print(f"   🏠 Aplicando filtro de tipo de propiedad: {property_type}")
                property_types = [pt.strip() for pt in property_type.split(',')]
                print(f"   🏠 Tipos parseados: {property_types}")
                type_conditions = []
                
                for prop_type in property_types:
                    property_type_lower = prop_type.lower()
                    if property_type_lower == "apartamento":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% apartamento %"),
                                Property.title.ilike("apartamento %"),
                                Property.title.ilike("% apartamento"),
                                Property.title.ilike("% apto %"),
                                Property.title.ilike("apto %"),
                                Property.title.ilike("% apto"),
                                Property.title.ilike("apartamento"),
                                Property.title.ilike("apto")
                            ),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "casa":
                        type_conditions.append(and_(
                            or_(
                                Property.title.ilike("% casa %"),
                                Property.title.ilike("casa %"),
                                Property.title.ilike("% casa"),
                                Property.title.ilike("casa")
                            ),
                            ~Property.title.ilike("%apartamento%"),
                            ~Property.title.ilike("%apto%"),
                            ~Property.title.ilike("%bodega%"),
                            ~Property.title.ilike("%local%"),
                            ~Property.title.ilike("%oficina%")
                        ))
                    elif property_type_lower == "oficina":
                        type_conditions.append(or_(
                            Property.title.ilike("% oficina %"),
                            Property.title.ilike("oficina %"),
                            Property.title.ilike("% oficina"),
                            Property.title.ilike("oficina")
                        ))
                    elif property_type_lower == "local":
                        type_conditions.append(or_(
                            Property.title.ilike("% local %"),
                            Property.title.ilike("local %"),
                            Property.title.ilike("% local"),
                            Property.title.ilike("local")
                        ))
                    elif property_type_lower == "bodega":
                        type_conditions.append(or_(
                            Property.title.ilike("% bodega %"),
                            Property.title.ilike("bodega %"),
                            Property.title.ilike("% bodega"),
                            Property.title.ilike("bodega")
                        ))
                    elif property_type_lower == "lote":
                        type_conditions.append(or_(
                            Property.title.ilike("% lote %"),
                            Property.title.ilike("lote %"),
                            Property.title.ilike("% lote"),
                            Property.title.ilike("lote")
                        ))
                    elif property_type_lower == "finca":
                        type_conditions.append(or_(
                            Property.title.ilike("% finca %"),
                            Property.title.ilike("finca %"),
                            Property.title.ilike("% finca"),
                            Property.title.ilike("finca")
                        ))
                
                if type_conditions:
                    print(f"   ✅ Agregando {len(type_conditions)} condiciones de tipo al filtro")
                    filters.append(or_(*type_conditions))
                else:
                    print(f"   ⚠️ No se generaron condiciones de tipo (lista vacía)")
            else:
                print(f"   ℹ️ No se aplicó filtro de tipo de propiedad (property_type is None or empty)")
            
            # Filtrar por fechas de actualización
            if updated_date_from:
                from datetime import datetime
                date_from = datetime.strptime(updated_date_from, '%Y-%m-%d').date()
                filters.append(Property.last_update >= date_from)
            
            if updated_date_to:
                from datetime import datetime
                date_to = datetime.strptime(updated_date_to, '%Y-%m-%d').date()
                filters.append(Property.last_update <= date_to)
            
            query = select(Property).where(and_(*filters))
            properties = session.exec(query).all()
            
            # Agrupar propiedades por coordenadas para retornarlas al frontend
            # El frontend usará Google Maps DDS para identificar a qué zona pertenece cada propiedad
            properties_data = []
            for prop in properties:
                properties_data.append({
                    'id': prop.fr_property_id,
                    'latitude': prop.latitude,
                    'longitude': prop.longitude,
                    'price': prop.price,
                    'offer': prop.offer,
                    'area': prop.area,
                    'rooms': prop.rooms,
                    'city_id': prop.city_id,
                    'location_main': prop.location_main,
                    'stratum': prop.stratum,
                    'title': prop.title,
                    'last_update': prop.last_update.isoformat() if prop.last_update else None
                })
            
            # Estadísticas generales
            total_properties = len(properties_data)
            properties_for_sale = [p for p in properties_data if p['offer'] == 'sell']
            properties_for_rent = [p for p in properties_data if p['offer'] == 'rent']
            
            return {
                'status': 'success',
                'boundary_type': boundary_type,
                'data': {
                    'properties': properties_data,
                    'summary': {
                        'total': total_properties,
                        'for_sale': len(properties_for_sale),
                        'for_rent': len(properties_for_rent)
                    }
                }
            }
            
    except Exception as e:
        print(f"Error getting properties by zone: {e}")
        return {
            'status': 'error',
            'message': str(e),
            'data': None
        }

@app.post("/api/property-valuation")
async def property_valuation(request: PropertyValuationRequest):
    """Endpoint para realizar avalúos de propiedades usando modelos ML"""
    try:
        # Importar dependencias ML solo cuando se necesiten
        import pandas as pd
        import numpy as np
        import joblib
        import lightgbm as lgb
        from catboost import CatBoostRegressor
        
        # Convertir request a diccionario
        property_data = request.dict()
        
        # Preprocesar datos para los modelos ML
        processed_data = property_data.copy()
        
        # Mantener valores categóricos como strings para LightGBM
        processed_data['is_new'] = property_data['is_new']  # "yes" o "no"
        processed_data['city_id'] = property_data['city_id']  # string
        
        # El modelo LightGBM fue entrenado con el formato original del frontend
        # Usar age_bucket tal como viene del frontend: "1-8", "9-15", etc.
        processed_data['age_bucket'] = property_data['age_bucket']
        
        # Rutas de archivos
        metadata_path = "/app/ml_models/metadata.json"
        rent_model_path = "/app/ml_models/model_rent_lightgbm.txt"
        sell_model_path = "/app/ml_models/model_sell_catboost.cbm"
        
        # Leer metadata para obtener el orden correcto de features
        import json
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)
        
        # Obtener el orden de features desde metadata
        rent_features = metadata['rent']['lightgbm']['features']
        sell_features = metadata['sell']['catboost']['features']
        
        results = {}
        
        # Cargar y predecir con modelo de renta (LightGBM original)
        try:
            if os.path.exists(rent_model_path):
                import lightgbm as lgb
                # Cargar modelo usando model_str para compatibilidad
                with open(rent_model_path, 'r') as f:
                    model_str = f.read()
                rent_model = lgb.Booster(model_str=model_str)
                
                # Crear DataFrame con las variables categóricas correctas
                df_rent = pd.DataFrame([processed_data])[rent_features]
                
                # Convertir variables categóricas a pandas categorical dtype
                categorical_features = ["is_new", "age_bucket", "city_id"]
                for cat_feat in categorical_features:
                    if cat_feat in df_rent.columns:
                        df_rent[cat_feat] = pd.Categorical(df_rent[cat_feat].astype(str))
                
                # Predecir con el modelo LightGBM
                rent_prediction = rent_model.predict(df_rent)[0]
                rent_price = float(np.expm1(rent_prediction))
                results['rent_price_per_sqm'] = round(rent_price, 2)
            else:
                results['rent_price_per_sqm'] = None
                results['rent_error'] = "Modelo de renta no encontrado"
        except Exception as e:
            results['rent_price_per_sqm'] = None
            results['rent_error'] = f"Error cargando modelo de renta: {str(e)}"
        
        # Cargar y predecir con modelo de venta (CatBoost)
        try:
            if os.path.exists(sell_model_path):
                sell_model = CatBoostRegressor()
                sell_model.load_model(sell_model_path)
                df_sell = pd.DataFrame([processed_data])[sell_features]
                sell_prediction = sell_model.predict(df_sell)[0]
                sell_price = float(np.expm1(sell_prediction))
                results['sell_price_per_sqm'] = round(sell_price, 2)
            else:
                results['sell_price_per_sqm'] = None
                results['sell_error'] = "Modelo de venta no encontrado"
        except Exception as e:
            results['sell_price_per_sqm'] = None
            results['sell_error'] = f"Error cargando modelo de venta: {str(e)}"
        
        # Calcular precios totales si hay predicciones exitosas
        if results.get('rent_price_per_sqm'):
            results['total_rent_price'] = round(results['rent_price_per_sqm'] * property_data['area'], 2)
        
        if results.get('sell_price_per_sqm'):
            results['total_sell_price'] = round(results['sell_price_per_sqm'] * property_data['area'], 2)
        
        return {
            'status': 'success',
            'message': 'Avalúo realizado exitosamente',
            'data': {
                'property_info': property_data,
                'valuation_results': results
            }
        }
        
    except Exception as e:
        print(f"Error en avalúo: {e}")
        return {
            'status': 'error',
            'message': f'Error realizando avalúo: {str(e)}',
            'data': None
        }

@app.post("/api/save-valuation")
async def save_valuation(request: SaveValuationRequest):
    """Endpoint para guardar o actualizar avalúos en la base de datos"""
    try:
        from sqlmodel import Session, select
        from config.db_connection import engine
        from models.valuation import Valuation
        from datetime import datetime
        from sqlalchemy.exc import IntegrityError
        
        with Session(engine) as session:
            # Verificar si ya existe un avalúo con ese nombre
            existing_valuation = session.exec(
                select(Valuation).where(Valuation.valuation_name == request.valuation_name)
            ).first()
            
            if existing_valuation:
                # Si existe, verificar si hay cambios en el precio final
                if existing_valuation.final_price != request.final_price:
                    # Actualizar el registro existente
                    existing_valuation.area = request.area
                    existing_valuation.property_type = request.property_type
                    existing_valuation.rooms = request.rooms
                    existing_valuation.baths = request.baths
                    existing_valuation.garages = request.garages
                    existing_valuation.stratum = request.stratum
                    existing_valuation.antiquity = request.antiquity
                    existing_valuation.latitude = request.latitude
                    existing_valuation.longitude = request.longitude
                    existing_valuation.capitalization_rate = request.capitalization_rate
                    existing_valuation.sell_price_per_sqm = request.sell_price_per_sqm
                    existing_valuation.rent_price_per_sqm = request.rent_price_per_sqm
                    existing_valuation.total_sell_price = request.total_sell_price
                    existing_valuation.total_rent_price = request.total_rent_price
                    existing_valuation.final_price = request.final_price
                    existing_valuation.updated_at = datetime.utcnow()
                    
                    session.commit()
                    session.refresh(existing_valuation)
                    
                    return {
                        'status': 'success',
                        'message': 'Avalúo actualizado exitosamente',
                        'valuation_id': existing_valuation.id,
                        'action': 'updated'
                    }
                else:
                    # El avalúo ya existe con el mismo precio final
                    return {
                        'status': 'error',
                        'message': f'Ya existe un avalúo con el nombre "{request.valuation_name}" y el mismo precio final. Use un nombre diferente o modifique el precio.',
                        'action': 'duplicate'
                    }
            
            # Si no existe, crear nuevo avalúo
            try:
                new_valuation = Valuation(
                    valuation_name=request.valuation_name,
                    area=request.area,
                    property_type=request.property_type,
                    rooms=request.rooms,
                    baths=request.baths,
                    garages=request.garages,
                    stratum=request.stratum,
                    antiquity=request.antiquity,
                    latitude=request.latitude,
                    longitude=request.longitude,
                    capitalization_rate=request.capitalization_rate,
                    sell_price_per_sqm=request.sell_price_per_sqm,
                    rent_price_per_sqm=request.rent_price_per_sqm,
                    total_sell_price=request.total_sell_price,
                    total_rent_price=request.total_rent_price,
                    final_price=request.final_price,
                    created_at=datetime.utcnow()
                )
                
                session.add(new_valuation)
                session.commit()
                session.refresh(new_valuation)
                
                return {
                    'status': 'success',
                    'message': 'Avalúo guardado exitosamente',
                    'valuation_id': new_valuation.id,
                    'action': 'created'
                }
                
            except IntegrityError as ie:
                session.rollback()
                return {
                    'status': 'error',
                    'message': f'Ya existe un avalúo con el nombre "{request.valuation_name}". Use un nombre diferente.',
                    'action': 'duplicate'
                }
            
    except Exception as e:
        print(f"Error guardando avalúo: {e}")
        return {
            'status': 'error',
            'message': f'Error guardando avalúo: {str(e)}'
        }

@app.delete("/api/valuations/{valuation_id}")
async def delete_valuation(valuation_id: int):
    """Endpoint para eliminar un avalúo"""
    try:
        from sqlmodel import Session, select
        from config.db_connection import engine
        from models.valuation import Valuation
        
        with Session(engine) as session:
            # Buscar el avalúo
            valuation = session.exec(select(Valuation).where(Valuation.id == valuation_id)).first()
            
            if not valuation:
                return {
                    'status': 'error',
                    'message': 'Avalúo no encontrado'
                }
            
            # Eliminar el avalúo
            session.delete(valuation)
            session.commit()
            
            return {
                'status': 'success',
                'message': 'Avalúo eliminado exitosamente'
            }
            
    except Exception as e:
        print(f"Error eliminando avalúo: {e}")
        return {
            'status': 'error',
            'message': f'Error eliminando avalúo: {str(e)}'
        }

@app.get("/api/valuations")
async def get_valuations(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page")
):
    """Endpoint para obtener avalúos con paginación"""
    try:
        from sqlmodel import Session, select, func
        from config.db_connection import engine
        from models.valuation import Valuation
        
        with Session(engine) as session:
            # Contar total de avalúos
            count_query = select(func.count(Valuation.id))
            total_count = session.exec(count_query).first() or 0
            
            # Calcular offset
            offset = (page - 1) * limit
            
            # Obtener avalúos ordenados del más reciente al más antiguo
            valuations_query = (
                select(Valuation)
                .order_by(Valuation.created_at.desc())
                .offset(offset)
                .limit(limit)
            )
            valuations = session.exec(valuations_query).all()
            
            # Calcular información de paginación
            total_pages = (total_count + limit - 1) // limit
            has_next = page < total_pages
            has_prev = page > 1
            
            # Convertir a diccionarios para la respuesta
            valuations_data = []
            for valuation in valuations:
                valuations_data.append({
                    "id": valuation.id,
                    "valuation_name": valuation.valuation_name,
                    "area": valuation.area,
                    "property_type": valuation.property_type,
                    "rooms": valuation.rooms,
                    "baths": valuation.baths,
                    "garages": valuation.garages,
                    "stratum": valuation.stratum,
                    "antiquity": valuation.antiquity,
                    "latitude": valuation.latitude,
                    "longitude": valuation.longitude,
                    "capitalization_rate": valuation.capitalization_rate,
                    "sell_price_per_sqm": valuation.sell_price_per_sqm,
                    "rent_price_per_sqm": valuation.rent_price_per_sqm,
                    "total_sell_price": valuation.total_sell_price,
                    "total_rent_price": valuation.total_rent_price,
                    "final_price": valuation.final_price,
                    "created_at": valuation.created_at.isoformat(),
                    "updated_at": valuation.updated_at.isoformat() if valuation.updated_at else None
                })
            
            return {
                "valuations": valuations_data,
                "pagination": {
                    "page": page,
                    "limit": limit,
                    "total_count": total_count,
                    "total_pages": total_pages,
                    "has_next": has_next,
                    "has_prev": has_prev
                }
            }
            
    except Exception as e:
        print(f"Error obteniendo avalúos: {e}")
        return {
            "valuations": [],
            "pagination": {
                "page": 1,
                "limit": 10,
                "total_count": 0,
                "total_pages": 0,
                "has_next": False,
                "has_prev": False
            }
        }


# Google Sheets Integration Models
class PaymentPlanRequest(BaseModel):
    """Request model for payment plan data"""
    # Identificación
    valuation_name: str  # Nombre del avalúo para el archivo
    
    # Flujo Interno
    area: str
    commercial_value: str
    average_purchase_value: str
    asking_price: str
    user_down_payment: str
    program_months: str
    potential_down_payment: str  # Ahora será un porcentaje
    bank_mortgage_rate: str
    dupla_bank_rate: str
    
    # Para Usuario
    client_name: str
    address: str
    city: str
    country: str
    construction_year: str
    stratum: str
    apartment_type: str
    private_parking: str


class PaymentPlanResponse(BaseModel):
    """Response model for payment plan creation"""
    success: bool
    sheet_url: str = ""
    message: str = ""


# Google Sheets Endpoints
@app.post("/api/google-sheets", response_model=PaymentPlanResponse)
async def create_payment_plan_sheet(payment_plan_data: PaymentPlanRequest):
    """
    Create a new Google Sheets document with payment plan data
    """
    import requests
    from urllib.parse import urlencode
    
    try:
        # Get Apps Script URL from environment
        forms_url = os.getenv('GOOGLE_APPS_SCRIPT_URL')
        if not forms_url:
            raise HTTPException(
                status_code=500, 
                detail="Google Apps Script URL not configured"
            )
        
        # Validate that required fields are present
        if not payment_plan_data.client_name.strip():
            raise HTTPException(
                status_code=400,
                detail="Client name is required"
            )
        
        # Convert Pydantic model to dictionary
        data_dict = payment_plan_data.model_dump()
        
        # Prepare URL parameters for GET request
        params = urlencode(data_dict)
        full_url = f"{forms_url}?{params}"
        
        # Debug: log the URL being called
        print(f"DEBUG: Calling Apps Script URL: {full_url[:200]}...")
        
        # Call Apps Script with requests (better redirect handling)
        response = requests.get(full_url, allow_redirects=True, timeout=30)
        print(f"DEBUG: Response status: {response.status_code}, URL after redirects: {response.url}")
        if response.status_code == 200:
            try:
                result = response.json()
                if result.get('success'):
                    return PaymentPlanResponse(
                        success=True,
                        sheet_url=result.get('sheet_url', ''),
                        message=result.get('message', 'Plan de pagos creado exitosamente')
                    )
                else:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Error en Apps Script: {result.get('error', 'Unknown error')}"
                    )
            except Exception as json_error:
                raise HTTPException(
                    status_code=500,
                    detail=f"Error parsing Apps Script response: {str(json_error)}"
                )
        else:
            raise HTTPException(
                status_code=500,
                detail=f"Error calling Apps Script: {response.status_code} - {response.text[:200]}"
            )
    
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    
    except Exception as e:
        print(f"Unexpected error in create_payment_plan_sheet: {e}")
        raise HTTPException(
            status_code=500,
            detail="Error al crear el plan de pagos en Google Sheets"
        )



@app.get("/")
async def root():
    return {"message": "Dashboard API is running", "timestamp": get_local_now()}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)